from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from io import BytesIO
import base64
import json
import inspect
import mimetypes
import re
from urllib.parse import quote_plus
from typing import Any

import pandas as pd
import streamlit as st

from database import (
    OFFER_DIR,
    UPLOAD_DIR,
    add_audit_log,
    add_document,
    add_offer_version,
    add_snapshot,
    get_audit_logs,
    get_benefit_records,
    get_candidate,
    get_career_records,
    get_compensation,
    get_document,
    get_documents,
    get_offer_versions,
    get_snapshots,
    get_setting,
    set_setting,
    get_revenue_rules,
    get_pay_band_reference,
    replace_pay_band_reference,
    get_pay_band_row,
    PAY_BAND_COLUMNS,
    replace_revenue_rules,
    get_company_mapping,
    upsert_company_mapping,
    list_company_mappings,
    delete_company_mapping,
    init_db,
    list_candidates,
    next_offer_version,
    replace_benefit_records,
    replace_career_records,
    update_document_analysis,
    upsert_candidate,
    upsert_compensation,
)
from ocr_service import (
    extract_text_from_file,
    parse_health_insurance_career,
    health_parse_diagnostics,
    tesseract_status,
)
from offer_letter import build_offer_ppt
from dart_service import (
    apply_revenue_rule,
    download_corp_codes,
    fetch_consolidated_revenue,
    find_company_match,
    load_corp_codes,
    normalize_company_name,
    normalize_corp_code,
)


OCR_PASSWORD_SUPPORTED = "password" in inspect.signature(extract_text_from_file).parameters
OCR_HEALTH_HINT_SUPPORTED = "health_document" in inspect.signature(extract_text_from_file).parameters


APP_TITLE = "경력직 처우검토 및 오퍼레터 관리시스템"
COMPANY_NAME = "우리회사"
HR_DEPARTMENT = "인사팀"

EMPLOYMENT_TYPES = ["정규직", "계약직", "파견직"]
JOB_GROUP_OPTIONS = ["일반직", "연구직", "별정직", "기술직"]
WORK_LOCATION_OPTIONS = ["인천", "당진", "포항", "순천", "판교"]
ENTRY_TYPE_OPTIONS = ["신입입사", "경력입사"]
CANDIDATE_STATUSES = ["검토중", "처우협의중", "오퍼승인", "오퍼발송", "입사확정", "종결"]

CAREER_RATE_GUIDE = {
    "검토필요": 0,
    "확인필요": 0,
    "매출기준": 0,
    "수동확정": 0,
    "미인정": 0,
}

BENEFIT_CATALOG_DEFAULT = [
    {"사용": True, "구분": "근무·휴가", "복리후생": "선택적 근로시간제", "설명": "의무 근무 시간(10~15시)을 제외한 선택 가능 시간대에서 출·퇴근 시간을 유연하게 활용할 수 있습니다."},
    {"사용": True, "구분": "근무·휴가", "복리후생": "하기 휴가", "설명": "연간 5일을 부여합니다. 7월 1일 이전 입사자에 한해 적용되며, 분할 사용이 가능합니다."},
    {"사용": True, "구분": "근무·휴가", "복리후생": "경조사 휴가/지원", "설명": "경조사별 최대 10일까지 휴가를 부여하며, 경조사 내용에 따라 경조금·경조화환 등 지원항목이 달라질 수 있습니다."},
    {"사용": True, "구분": "근무·휴가", "복리후생": "육아휴직, 가족돌봄휴가", "설명": "육아휴직은 최대 2년까지 지원하며, 가족돌봄휴가 등 가족친화 제도를 운영합니다."},
    {"사용": True, "구분": "근무·휴가", "복리후생": "유급 휴일", "설명": "취업규칙상 유급휴일을 부여합니다. 건강검진일, 창립기념일, 연시·설날·추석 법정휴일 추가 1일 등이 포함될 수 있습니다."},
    {"사용": True, "구분": "의료·생활 편의", "복리후생": "의료비, 입원비, 건강검진비", "설명": "본인·배우자·부양가족 의료비 및 입원비를 지원합니다. 입원비는 3,000만원 한도, 본인·배우자·자녀·형제 등 검진비는 30만원 한도로 운영될 수 있습니다."},
    {"사용": True, "구분": "의료·생활 편의", "복리후생": "호텔/콘도", "설명": "전국 해비치·한화·대명·리솜리조트 등 제휴 휴양시설 이용을 지원합니다."},
    {"사용": True, "구분": "의료·생활 편의", "복리후생": "출퇴근 셔틀버스", "설명": "서울·경기 주요 지역 출퇴근 셔틀버스를 운영합니다. 실제 노선과 이용 가능 여부는 근무지별 운영기준에 따릅니다."},
    {"사용": True, "구분": "건강", "복리후생": "단체상해보험", "설명": "재직 중 발생할 수 있는 질병·상해 위험에 대비할 수 있도록 회사 기준에 따라 단체보험 가입을 지원합니다."},
    {"사용": True, "구분": "주거", "복리후생": "주택자금 지원", "설명": "회사 기준과 자격요건에 따라 주택 구입·임차 관련 자금 또는 금융지원을 제공합니다."},
    {"사용": True, "구분": "주거", "복리후생": "사택·기숙사 지원", "설명": "근무지 및 입주요건에 따라 사택 또는 기숙사 등 주거 편의를 지원합니다."},
    {"사용": True, "구분": "생활", "복리후생": "생활안정자금 지원", "설명": "회사 기준에 따라 임직원의 생활안정을 위한 대출 또는 금융지원 제도를 운영합니다."},
    {"사용": True, "구분": "가족", "복리후생": "자녀 학자금 지원", "설명": "자녀의 교육비 부담 경감을 위해 회사 기준에 따른 학자금 또는 교육비를 지원합니다."},
    {"사용": True, "구분": "가족", "복리후생": "임신·출산 지원", "설명": "임신·출산과 관련한 휴가, 선물 또는 기타 가족친화 지원을 회사 기준에 따라 제공합니다."},
    {"사용": True, "구분": "가족", "복리후생": "사내 어린이집", "설명": "운영 사업장 또는 이용요건에 따라 임직원 자녀를 위한 보육시설 이용을 지원합니다."},
    {"사용": True, "구분": "생활", "복리후생": "차량구입 지원", "설명": "현대자동차·기아 차량 구입 시 회사 및 그룹 기준에 따른 임직원 할인 또는 구매지원을 제공합니다."},
    {"사용": True, "구분": "생활", "복리후생": "명절·기념일 지원", "설명": "명절, 생일, 창립기념일 등 회사가 정한 시점에 선물 또는 기념품 등을 지원합니다."},
    {"사용": True, "구분": "근무환경", "복리후생": "사내식당·식사 지원", "설명": "사업장별 운영기준에 따라 사내식당 또는 식사 관련 편의를 제공합니다."},
    {"사용": True, "구분": "근무환경", "복리후생": "휴게·편의시설", "설명": "사업장 여건에 따라 휴게공간, 카페테리아 등 임직원 편의시설을 운영합니다."},
    {"사용": True, "구분": "문화", "복리후생": "사내 동호회 지원", "설명": "임직원의 건전한 여가활동과 교류를 위해 회사 기준에 따라 동호회 활동을 지원합니다."},
    {"사용": True, "구분": "포상", "복리후생": "장기근속 포상", "설명": "일정 근속연수를 충족한 임직원에게 회사 기준에 따라 포상 또는 휴가 등을 제공합니다."},
    {"사용": True, "구분": "포상", "복리후생": "우수사원 포상", "설명": "회사 발전과 성과에 기여한 임직원을 대상으로 회사 기준에 따른 포상제도를 운영합니다."},
    {"사용": True, "구분": "교육", "복리후생": "직무·어학 교육 지원", "설명": "직무역량 및 글로벌 역량 향상을 위한 사내외 교육, 어학과정 등을 회사 기준에 따라 지원합니다."},
    {"사용": True, "구분": "교육", "복리후생": "자격증 취득 지원", "설명": "직무 관련 자격 취득을 장려하기 위해 회사 기준에 따른 응시·취득 지원제도를 운영합니다."},
]

BENEFIT_COLUMNS = ["오퍼레터 포함", "구분", "복리후생", "적용여부", "설명"]
BENEFIT_CATALOG_SETTING = "BENEFIT_CATALOG_V36"
BENEFIT_PROFILES_SETTING = "BENEFIT_PROFILES_V36"

CAREER_COLUMNS = [
    "회사명", "입사일", "퇴사일", "직무/직책",
    "표준회사명", "DART고유번호", "매출연도", "연결매출(억원)",
    "매출조회상태", "매출기준인정률(%)", "공시접수번호",
    "외부자료출처", "외부자료URL", "재무기준", "외부증빙파일",
    "수동확정", "수동조정사유", "인정구분", "인정률(%)",
    "출처", "출처파일", "비고",
]

DART_CORP_CACHE = Path(__file__).resolve().parent / "data" / "dart_corp_codes.csv"

GRADE_OPTIONS = ["G1", "G2", "G3", "G4", "R1", "R2", "R3", "R4"]
CASH_BENEFIT_FIELDS = [
    ("직책수당", "c_cash_job_allowance"),
    ("가족수당(본인)", "c_cash_family_allowance"),
    ("차량보조금", "c_cash_vehicle_subsidy"),
    ("자기계발지원금", "c_cash_self_development"),
    ("귀향여비", "c_cash_homecoming_travel"),
    ("복지포인트", "c_cash_welfare_points"),
    ("선택형 복지포인트", "c_cash_flexible_welfare_points"),
    ("체력단련비", "c_cash_fitness"),
    ("개인연금", "c_cash_pension"),
]

def pay_grade_group(grade: str, role: str = "일반") -> str:
    grade = str(grade or "G1").upper()
    level = grade[-1] if grade and grade[-1] in "1234" else "1"
    base = f"G{level}/R{level}"
    if level == "4" and role in {"팀장", "실장"}:
        return f"{base}({role})"
    return base

def parse_pay_band_csv(uploaded_file) -> pd.DataFrame:
    raw = uploaded_file.getvalue()
    last_error = None
    for encoding in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            df = pd.read_csv(BytesIO(raw), encoding=encoding)
            break
        except Exception as exc:
            last_error = exc
    else:
        raise ValueError(f"CSV 파일을 읽지 못했습니다: {last_error}")
    df.columns = [str(col).strip() for col in df.columns]
    if "현금성지금소계" in df.columns and "현금성지급소계" not in df.columns:
        df = df.rename(columns={"현금성지금소계": "현금성지급소계"})
    missing = [col for col in PAY_BAND_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError("필수 열이 없습니다: " + ", ".join(missing))
    return df[PAY_BAND_COLUMNS].copy()

def normalize_pay_band_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "현금성지금소계" in out.columns and "현금성지급소계" not in out.columns:
        out = out.rename(columns={"현금성지금소계": "현금성지급소계"})
    for col in PAY_BAND_COLUMNS:
        if col not in out.columns:
            out[col] = "" if col in {"직급", "BAND"} else 0.0
    out = out[PAY_BAND_COLUMNS]
    for col in PAY_BAND_COLUMNS[2:]:
        out[col] = pd.to_numeric(out[col].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0.0)
    out["직급"] = out["직급"].fillna("").astype(str).str.strip()
    out["BAND"] = out["BAND"].fillna("").astype(str).str.strip()
    return out

def format_won(value: float | int) -> str:
    return f"{float(value or 0):,.0f}원"


def _parse_money_text(value: Any) -> int:
    text = str(value or "").replace(",", "").replace("원", "").strip()
    if not text:
        return 0
    text = re.sub(r"[^0-9]", "", text)
    return int(text or 0)


def money_input(label: str, key: str, step: int = 100000, help: str | None = None) -> int:
    # st.number_input은 버전에 따라 천단위 쉼표 입력 UX가 달라질 수 있어
    # 금액은 문자열 위젯으로 보여주고 내부 값은 원 단위 정수로 유지합니다.
    try:
        numeric = max(int(round(float(st.session_state.get(key, 0) or 0))), 0)
    except Exception:
        numeric = 0
    st.session_state[key] = numeric

    display_key = f"__money_display_{key}"
    last_key = f"__money_last_{key}"
    if display_key not in st.session_state or st.session_state.get(last_key) != numeric:
        st.session_state[display_key] = f"{numeric:,}"
        st.session_state[last_key] = numeric

    def _sync_money() -> None:
        parsed = _parse_money_text(st.session_state.get(display_key, "0"))
        st.session_state[key] = parsed
        # last_key는 일부러 갱신하지 않습니다. 다음 rerun 시작 시 쉼표 형식으로 재표시됩니다.

    st.text_input(
        label, key=display_key, help=help, on_change=_sync_money,
        placeholder="0",
    )
    return int(st.session_state.get(key, 0) or 0)


def format_percent(value: float | int) -> str:
    return f"{float(value or 0):+,.1f}%"


def safe_filename(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", text.strip()) or "후보자"


def days_to_ymd(total_days: float | int) -> tuple[int, int, int]:
    days = max(int(round(float(total_days or 0))), 0)
    years = days // 365
    remain = days % 365
    months = remain // 30
    return years, months, remain % 30


def period_text(total_days: float | int) -> str:
    years, months, days = days_to_ymd(total_days)
    return f"{years}년 {months}개월 {days}일"


def normalize_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return parsed.date().isoformat()


def ensure_career_columns(df: pd.DataFrame | None) -> pd.DataFrame:
    result = pd.DataFrame() if df is None else df.copy()
    defaults = {
        "회사명": "", "입사일": pd.NaT, "퇴사일": pd.NaT, "직무/직책": "",
        "표준회사명": "", "DART고유번호": "", "매출연도": pd.NA,
        "연결매출(억원)": pd.NA, "매출조회상태": "미조회",
        "매출기준인정률(%)": pd.NA, "공시접수번호": "",
        "외부자료출처": "", "외부자료URL": "", "재무기준": "", "외부증빙파일": "",
        "수동확정": False, "수동조정사유": "",
        "인정구분": "확인필요", "인정률(%)": 0,
        "출처": "수기입력", "출처파일": "", "비고": "",
    }
    for column in CAREER_COLUMNS:
        if column not in result.columns:
            result[column] = defaults[column]
    result["수동확정"] = result["수동확정"].fillna(False).astype(bool)
    return result[CAREER_COLUMNS]


def default_career_df() -> pd.DataFrame:
    return ensure_career_columns(pd.DataFrame())


def normalize_legacy_employment_type(value: str) -> str:
    value = str(value or "").strip()
    if value in EMPLOYMENT_TYPES:
        return value
    if value in {"기간제 계약직", "촉탁계약직", "전문계약직"}:
        return "계약직"
    return "정규직"


def normalize_work_location(value: str) -> str:
    value = str(value or "").strip()
    for location in WORK_LOCATION_OPTIONS:
        if location in value:
            return location
    return "당진"


def benefit_catalog_df() -> pd.DataFrame:
    raw = get_setting(BENEFIT_CATALOG_SETTING, "")
    if raw:
        try:
            loaded = json.loads(raw)
            df = pd.DataFrame(loaded)
        except Exception:
            df = pd.DataFrame(BENEFIT_CATALOG_DEFAULT)
    else:
        df = pd.DataFrame(BENEFIT_CATALOG_DEFAULT)
    for col, default in {"사용": True, "구분": "기타", "복리후생": "", "설명": ""}.items():
        if col not in df.columns:
            df[col] = default
    df = df[["사용", "구분", "복리후생", "설명"]].copy()
    df["사용"] = df["사용"].fillna(True).astype(bool)
    for col in ["구분", "복리후생", "설명"]:
        df[col] = df[col].fillna("").astype(str)
    return df


def save_benefit_catalog(df: pd.DataFrame) -> None:
    cleaned = df.copy()
    cleaned = cleaned[cleaned["복리후생"].fillna("").astype(str).str.strip() != ""].copy()
    cleaned["복리후생"] = cleaned["복리후생"].astype(str).str.strip()
    cleaned = cleaned.drop_duplicates(subset=["복리후생"], keep="last")
    set_setting(BENEFIT_CATALOG_SETTING, json.dumps(cleaned.to_dict("records"), ensure_ascii=False))


def benefit_profiles() -> dict[str, list[str]]:
    raw = get_setting(BENEFIT_PROFILES_SETTING, "")
    if not raw:
        return {}
    try:
        value = json.loads(raw)
        return value if isinstance(value, dict) else {}
    except Exception:
        return {}


def save_benefit_profiles(profiles: dict[str, list[str]]) -> None:
    set_setting(BENEFIT_PROFILES_SETTING, json.dumps(profiles, ensure_ascii=False))


def benefit_profile_key(job_group: str, contract_type: str, work_location: str, entry_type: str, grade: str) -> str:
    return "|".join([job_group, contract_type, work_location, entry_type, grade])


def benefit_profile_df(job_group: str, contract_type: str, work_location: str, entry_type: str, grade: str) -> pd.DataFrame:
    key = benefit_profile_key(job_group, contract_type, work_location, entry_type, grade)
    selected = set(benefit_profiles().get(key, []))
    catalog = benefit_catalog_df()
    catalog = catalog[(catalog["사용"] == True) & (catalog["복리후생"].isin(selected))].copy()  # noqa: E712
    rows = [{
        "오퍼레터 포함": True,
        "구분": row["구분"],
        "복리후생": row["복리후생"],
        "적용여부": "적용",
        "설명": row["설명"],
    } for row in catalog.to_dict("records")]
    return pd.DataFrame(rows, columns=BENEFIT_COLUMNS)


def empty_benefit_df() -> pd.DataFrame:
    return pd.DataFrame(columns=BENEFIT_COLUMNS)


def calculate_career(career_df: pd.DataFrame, calculation_date: date) -> dict[str, Any]:
    df = ensure_career_columns(career_df)
    if df.empty:
        return {
            "detail": pd.DataFrame(), "raw_days": 0, "recognized_days": 0,
            "pending_days": 0, "errors": [],
        }

    df["회사명"] = df["회사명"].fillna("").astype(str).str.strip()
    df = df[df["회사명"] != ""].copy()
    if df.empty:
        return {
            "detail": pd.DataFrame(), "raw_days": 0, "recognized_days": 0,
            "pending_days": 0, "errors": [],
        }

    df["입사일"] = pd.to_datetime(df["입사일"], errors="coerce")
    df["퇴사일"] = pd.to_datetime(df["퇴사일"], errors="coerce")
    df["인정률(%)"] = pd.to_numeric(df["인정률(%)"], errors="coerce").fillna(0).clip(0, 100)

    errors: list[str] = []
    raw_dates: set[date] = set()
    pending_dates: set[date] = set()
    confirmed_date_rates: dict[date, float] = {}
    detail_rows: list[dict[str, Any]] = []

    for index, row in df.iterrows():
        start_date_value = row["입사일"]
        end_date_value = row["퇴사일"]
        company = row["회사명"]

        if pd.isna(start_date_value):
            errors.append(f"{index + 1}행({company}): 입사일이 없습니다.")
            continue
        if pd.isna(end_date_value):
            end_date_value = pd.Timestamp(calculation_date)
        if end_date_value < start_date_value:
            errors.append(f"{index + 1}행({company}): 퇴사일이 입사일보다 빠릅니다.")
            continue

        start_date = start_date_value.date()
        end_date = min(end_date_value.date(), calculation_date)
        rate = float(row["인정률(%)"])
        manual = bool(row.get("수동확정", False))
        revenue_status = str(row.get("매출조회상태", "") or "")
        confirmed = manual or revenue_status == "조회완료" or str(row.get("인정구분", "")) == "미인정"
        total_days = (end_date - start_date).days + 1

        for day in pd.date_range(start=start_date, end=end_date, freq="D"):
            key = day.date()
            raw_dates.add(key)
            if confirmed:
                confirmed_date_rates[key] = max(confirmed_date_rates.get(key, 0), rate)
            else:
                pending_dates.add(key)

        detail_rows.append({
            "회사명": company,
            "표준회사명": str(row.get("표준회사명", "") or ""),
            "기간": f"{start_date:%Y-%m-%d} ~ {end_date:%Y-%m-%d}",
            "연결매출": (
                f"{float(row['연결매출(억원)']):,.0f}억원"
                if pd.notna(row.get("연결매출(억원)")) and str(row.get("연결매출(억원)")) != ""
                else "확인필요"
            ),
            "매출조회상태": revenue_status or "확인필요",
            "인정구분": str(row.get("인정구분", "") or ""),
            "인정률": f"{rate:.0f}%" if confirmed else "확인필요",
            "실경력": period_text(total_days),
            "행별 인정경력": period_text(total_days * rate / 100) if confirmed else "확인필요",
            "비고": str(row.get("비고", "") or ""),
        })

    # 이미 확정된 날짜는 확인필요 기간에서 제거하여 지표 중복을 피합니다.
    pending_only_dates = pending_dates - set(confirmed_date_rates.keys())
    return {
        "detail": pd.DataFrame(detail_rows),
        "raw_days": len(raw_dates),
        "recognized_days": sum(rate / 100 for rate in confirmed_date_rates.values()),
        "pending_days": len(pending_only_dates),
        "errors": errors,
    }

def ensure_session_defaults() -> None:
    defaults = {
        "current_candidate_id": None,
        "loaded_candidate_id": None,
        "candidate_code": "",
        "career_df": default_career_df(),
        "career_editor_version": 0,
        "benefit_df": empty_benefit_df(),
        "benefit_editor_version": 0,
        "nav_page": "후보자 목록",
        "f_name": "",
        "f_email": "",
        "f_phone": "",
        "f_gender": "",
        "f_birth_date": None,
        "f_education": "",
        "f_major": "",
        "f_photo_path": "",
        "f_current_company": "",
        "f_department": "",
        "f_target_job": "",
        "f_work_location": "당진",
        "f_employment_type": "정규직",
        "f_job_group": "일반직",
        "f_entry_type": "경력입사",
        "f_expected_join_date": date.today() + timedelta(days=30),
        "f_status": "검토중",
        "f_notes": "",
        "c_current_contract_salary": 0,
        "c_current_total_salary": 0,
        "c_grade": "G1",
        "c_grade_role": "일반",
        "c_band": "",
        "c_last_band_key": "",
        "c_offer_base_salary": 0,
        "c_offer_performance_salary": 0,
        "c_offer_fixed_overtime": 0,
        "c_offer_incentive": 0,
        "c_cash_job_allowance": 0,
        "c_cash_family_allowance": 0,
        "c_cash_vehicle_subsidy": 0,
        "c_cash_self_development": 0,
        "c_cash_homecoming_travel": 0,
        "c_cash_welfare_points": 0,
        "c_cash_flexible_welfare_points": 0,
        "c_cash_fitness": 0,
        "c_cash_pension": 0,
        "c_sign_on_bonus": 0,
        "c_promotion_base_date": "",
        "offer_date": date.today(),
        "acceptance_deadline": date.today() + timedelta(days=7),
        "special_terms": "",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_candidate_state() -> None:
    keys_to_delete = [
        "current_candidate_id", "loaded_candidate_id", "candidate_code",
        "career_df", "benefit_df",
        "f_name", "f_email", "f_phone", "f_gender", "f_birth_date", "f_education", "f_major", "f_photo_path", "f_current_company",
        "f_department", "f_target_job", "f_work_location",
        "f_employment_type", "f_job_group", "f_entry_type", "f_expected_join_date", "f_status", "f_notes",
        "c_current_contract_salary", "c_current_total_salary",
        "c_grade", "c_grade_role", "c_band", "c_last_band_key",
        "c_offer_base_salary", "c_offer_performance_salary", "c_offer_fixed_overtime", "c_offer_incentive",
        "c_cash_job_allowance", "c_cash_family_allowance", "c_cash_vehicle_subsidy",
        "c_cash_self_development", "c_cash_homecoming_travel", "c_cash_welfare_points",
        "c_cash_flexible_welfare_points", "c_cash_fitness", "c_cash_pension", "c_sign_on_bonus", "c_promotion_base_date",
        "offer_date", "acceptance_deadline", "special_terms",
    ]
    for key in keys_to_delete:
        st.session_state.pop(key, None)
    st.session_state["career_editor_version"] = st.session_state.get("career_editor_version", 0) + 1
    st.session_state["benefit_editor_version"] = st.session_state.get("benefit_editor_version", 0) + 1
    ensure_session_defaults()


def load_candidate_state(candidate_id: int) -> None:
    candidate = get_candidate(candidate_id)
    if not candidate:
        st.error("후보자 정보를 찾을 수 없습니다.")
        return

    st.session_state.current_candidate_id = candidate_id
    st.session_state.loaded_candidate_id = candidate_id
    st.session_state.candidate_code = candidate["candidate_code"]
    st.session_state.f_name = candidate.get("name", "")
    st.session_state.f_email = candidate.get("email", "")
    st.session_state.f_phone = candidate.get("phone", "")
    st.session_state.f_gender = candidate.get("gender", "") or ""
    birth = pd.to_datetime(candidate.get("birth_date"), errors="coerce")
    st.session_state.f_birth_date = birth.date() if not pd.isna(birth) else None
    st.session_state.f_education = candidate.get("education", "") or ""
    st.session_state.f_major = candidate.get("major", "") or ""
    st.session_state.f_photo_path = candidate.get("photo_path", "") or ""
    st.session_state.f_current_company = candidate.get("current_company", "")
    st.session_state.f_department = candidate.get("department", "")
    st.session_state.f_target_job = candidate.get("target_job", "")
    st.session_state.f_work_location = normalize_work_location(candidate.get("work_location", ""))
    st.session_state.f_employment_type = normalize_legacy_employment_type(candidate.get("employment_type", "정규직"))
    st.session_state.f_job_group = candidate.get("job_group", "") if candidate.get("job_group") in JOB_GROUP_OPTIONS else "일반직"
    st.session_state.f_entry_type = candidate.get("entry_type", "") if candidate.get("entry_type") in ENTRY_TYPE_OPTIONS else "경력입사"
    expected = pd.to_datetime(candidate.get("expected_join_date"), errors="coerce")
    st.session_state.f_expected_join_date = (
        expected.date() if not pd.isna(expected) else date.today() + timedelta(days=30)
    )
    st.session_state.f_status = candidate.get("status", "검토중")
    st.session_state.f_notes = candidate.get("notes", "")

    careers = get_career_records(candidate_id)
    career_df = ensure_career_columns(pd.DataFrame(careers))
    for column in ["입사일", "퇴사일"]:
        if column in career_df.columns:
            career_df[column] = pd.to_datetime(career_df[column], errors="coerce")
    st.session_state.career_df = career_df
    st.session_state.career_editor_version += 1

    compensation = get_compensation(candidate_id) or {}
    old_current_contract = float(compensation.get("current_fixed_salary", 0) or 0)
    old_current_total = (
        float(compensation.get("current_fixed_salary", 0) or 0)
        + float(compensation.get("current_incentive", 0) or 0)
        + float(compensation.get("current_other", 0) or 0)
    )
    st.session_state.c_current_contract_salary = int(round(float(compensation.get("current_contract_salary", 0) or old_current_contract)))
    st.session_state.c_current_total_salary = int(round(float(compensation.get("current_total_salary", 0) or old_current_total)))
    selected_grade = str(compensation.get("selected_grade", "") or "G1")
    if selected_grade not in GRADE_OPTIONS:
        selected_grade = "G1"
    st.session_state.c_grade = selected_grade
    selected_group = str(compensation.get("selected_band", "") or "")
    if selected_group.startswith("G4/R4(팀장)"):
        st.session_state.c_grade_role = "팀장"
    elif selected_group.startswith("G4/R4(실장)"):
        st.session_state.c_grade_role = "실장"
    else:
        st.session_state.c_grade_role = "일반"
    st.session_state.c_band = selected_group.split("|", 1)[1] if "|" in selected_group else ""
    st.session_state.c_last_band_key = f"{pay_grade_group(st.session_state.c_grade, st.session_state.c_grade_role)}|{st.session_state.c_band}" if st.session_state.c_band else ""

    compensation_map = {
        "c_offer_base_salary": ("offer_base_salary", 0),
        "c_offer_performance_salary": ("offer_performance_salary", 0),
        "c_offer_fixed_overtime": ("offer_fixed_overtime", compensation.get("offer_fixed_allowance", 0)),
        "c_offer_incentive": ("offer_incentive", compensation.get("offer_target_incentive", 0)),
        "c_cash_job_allowance": ("cash_job_allowance", 0),
        "c_cash_family_allowance": ("cash_family_allowance", 0),
        "c_cash_vehicle_subsidy": ("cash_vehicle_subsidy", 0),
        "c_cash_self_development": ("cash_self_development", 0),
        "c_cash_homecoming_travel": ("cash_homecoming_travel", 0),
        "c_cash_welfare_points": ("cash_welfare_points", 0),
        "c_cash_flexible_welfare_points": ("cash_flexible_welfare_points", 0),
        "c_cash_fitness": ("cash_fitness", 0),
        "c_cash_pension": ("cash_pension", 0),
        "c_sign_on_bonus": ("sign_on_bonus", 0),
    }
    for session_key, (db_key, fallback) in compensation_map.items():
        st.session_state[session_key] = int(round(float(compensation.get(db_key, 0) or fallback or 0)))
    st.session_state.c_promotion_base_date = compensation.get("promotion_base_date", "") or ""

    benefits = get_benefit_records(candidate_id)
    st.session_state.benefit_df = (
        pd.DataFrame(benefits, columns=BENEFIT_COLUMNS)
        if benefits
        else benefit_profile_df(
            st.session_state.f_job_group, st.session_state.f_employment_type,
            st.session_state.f_work_location, st.session_state.f_entry_type, st.session_state.c_grade,
        )
    )
    st.session_state.benefit_editor_version += 1


def career_records_for_db(df: pd.DataFrame) -> list[dict[str, Any]]:
    records = []
    df = ensure_career_columns(df)
    for row in df.to_dict("records"):
        records.append({
            **row,
            "입사일": normalize_date(row.get("입사일")),
            "퇴사일": normalize_date(row.get("퇴사일")),
        })
    return records


def current_compensation_data() -> dict[str, Any]:
    grade_group = pay_grade_group(st.session_state.c_grade, st.session_state.c_grade_role)
    selected_band_key = f"{grade_group}|{st.session_state.c_band}" if st.session_state.c_band else ""
    return {
        "current_contract_salary": st.session_state.c_current_contract_salary,
        "current_total_salary": st.session_state.c_current_total_salary,
        "selected_grade": st.session_state.c_grade,
        "selected_band": selected_band_key,
        "offer_base_salary": st.session_state.c_offer_base_salary,
        "offer_performance_salary": st.session_state.c_offer_performance_salary,
        "offer_fixed_overtime": st.session_state.c_offer_fixed_overtime,
        "offer_incentive": st.session_state.c_offer_incentive,
        "cash_job_allowance": st.session_state.c_cash_job_allowance,
        "cash_family_allowance": st.session_state.c_cash_family_allowance,
        "cash_vehicle_subsidy": st.session_state.c_cash_vehicle_subsidy,
        "cash_self_development": st.session_state.c_cash_self_development,
        "cash_homecoming_travel": st.session_state.c_cash_homecoming_travel,
        "cash_welfare_points": st.session_state.c_cash_welfare_points,
        "cash_flexible_welfare_points": st.session_state.c_cash_flexible_welfare_points,
        "cash_fitness": st.session_state.c_cash_fitness,
        "cash_pension": st.session_state.c_cash_pension,
        "sign_on_bonus": st.session_state.c_sign_on_bonus,
        "promotion_base_date": st.session_state.c_promotion_base_date,
    }


def build_snapshot() -> dict[str, Any]:
    career_result = calculate_career(st.session_state.career_df, date.today())
    compensation = current_compensation_data()
    offered_contract = (
        compensation["offer_base_salary"]
        + compensation["offer_performance_salary"]
        + compensation["offer_fixed_overtime"]
    )
    cash_benefits = sum(compensation[key] for key in [
        "cash_job_allowance", "cash_family_allowance", "cash_vehicle_subsidy",
        "cash_self_development", "cash_homecoming_travel", "cash_welfare_points",
        "cash_flexible_welfare_points", "cash_fitness", "cash_pension",
    ])
    offered_total = offered_contract + compensation["offer_incentive"] + cash_benefits
    current_total = compensation["current_total_salary"]
    return {
        "candidate": {
            "id": st.session_state.current_candidate_id,
            "candidate_code": st.session_state.candidate_code,
            "name": st.session_state.f_name,
            "email": st.session_state.f_email,
            "phone": st.session_state.f_phone,
            "gender": st.session_state.f_gender,
            "birth_date": str(st.session_state.f_birth_date or ""),
            "education": st.session_state.f_education,
            "major": st.session_state.f_major,
            "photo_path": st.session_state.f_photo_path,
            "current_company": st.session_state.f_current_company,
            "department": st.session_state.f_department,
            "target_job": st.session_state.f_target_job,
            "work_location": st.session_state.f_work_location,
            "employment_type": st.session_state.f_employment_type,
            "job_group": st.session_state.f_job_group,
            "entry_type": st.session_state.f_entry_type,
            "expected_join_date": str(st.session_state.f_expected_join_date),
            "status": st.session_state.f_status,
            "notes": st.session_state.f_notes,
        },
        "careers": career_records_for_db(st.session_state.career_df),
        "career_result": {
            "raw_days": career_result["raw_days"],
            "recognized_days": career_result["recognized_days"],
            "recognized_text": period_text(career_result["recognized_days"]),
        },
        "compensation": compensation,
        "compensation_result": {
            "current_contract": compensation["current_contract_salary"],
            "offered_contract": offered_contract,
            "contract_difference": offered_contract - compensation["current_contract_salary"],
            "contract_difference_rate": ((offered_contract - compensation["current_contract_salary"]) / compensation["current_contract_salary"] * 100 if compensation["current_contract_salary"] else 0),
            "current_total": current_total,
            "cash_benefits": cash_benefits,
            "offered_total": offered_total,
            "difference": offered_total - current_total,
            "difference_rate": ((offered_total - current_total) / current_total * 100 if current_total else 0),
        },
        "benefits": st.session_state.benefit_df.to_dict("records"),
        "saved_at": datetime.now().isoformat(timespec="seconds"),
    }


def save_current_candidate(reason: str = "후보자 정보 저장", show_message: bool = True) -> int | None:
    name = st.session_state.f_name.strip()
    if not name:
        st.error("후보자 성명을 입력하세요.")
        return None

    candidate_id = upsert_candidate({
        "id": st.session_state.current_candidate_id,
        "candidate_code": st.session_state.candidate_code,
        "name": name,
        "email": st.session_state.f_email,
        "phone": st.session_state.f_phone,
        "gender": st.session_state.f_gender,
        "birth_date": st.session_state.f_birth_date.isoformat() if st.session_state.f_birth_date else "",
        "education": st.session_state.f_education,
        "major": st.session_state.f_major,
        "photo_path": st.session_state.f_photo_path,
        "current_company": st.session_state.f_current_company,
        "department": st.session_state.f_department,
        "target_job": st.session_state.f_target_job,
        "work_location": st.session_state.f_work_location,
        "employment_type": st.session_state.f_employment_type,
        "job_group": st.session_state.f_job_group,
        "entry_type": st.session_state.f_entry_type,
        "expected_join_date": st.session_state.f_expected_join_date.isoformat(),
        "status": st.session_state.f_status,
        "notes": st.session_state.f_notes,
    })
    st.session_state.current_candidate_id = candidate_id

    candidate = get_candidate(candidate_id)
    st.session_state.candidate_code = candidate["candidate_code"] if candidate else ""

    replace_career_records(candidate_id, career_records_for_db(st.session_state.career_df))
    upsert_compensation(candidate_id, current_compensation_data())
    replace_benefit_records(
        candidate_id,
        st.session_state.benefit_df.to_dict("records"),
    )

    snapshot = build_snapshot()
    add_snapshot(candidate_id, snapshot, reason)
    add_audit_log(candidate_id, reason, {
        "status": st.session_state.f_status,
        "recognized_career": snapshot["career_result"]["recognized_text"],
        "offered_total": snapshot["compensation_result"]["offered_total"],
    })

    if show_message:
        st.success(f"저장 완료: {st.session_state.candidate_code} / {name}")
    return candidate_id


def save_uploaded_file(candidate_code: str, original_name: str, content: bytes) -> Path:
    candidate_dir = UPLOAD_DIR / safe_filename(candidate_code)
    candidate_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    target = candidate_dir / f"{timestamp}_{safe_filename(original_name)}"
    target.write_bytes(content)
    return target


def go_to_new_candidate() -> None:
    """신규 후보자 화면으로 이동합니다.

    Streamlit 위젯의 key 값은 위젯 생성 후 같은 실행 사이클에서
    직접 변경할 수 없으므로 버튼 콜백에서 먼저 변경합니다.
    """
    reset_candidate_state()
    st.session_state["nav_page"] = "후보자 검토"


def open_candidate_and_go(candidate_id: int) -> None:
    """선택한 후보자를 불러온 뒤 검토 화면으로 이동합니다."""
    load_candidate_state(int(candidate_id))
    st.session_state["nav_page"] = "후보자 검토"


def reload_current_candidate() -> None:
    """현재 후보자의 DB 저장값을 다시 불러옵니다."""
    candidate_id = st.session_state.get("current_candidate_id")
    if candidate_id:
        load_candidate_state(int(candidate_id))


def compensation_result() -> dict[str, Any]:
    data = current_compensation_data()
    offered_contract = data["offer_base_salary"] + data["offer_performance_salary"] + data["offer_fixed_overtime"]
    cash_benefits = sum(data[key] for key in [
        "cash_job_allowance", "cash_family_allowance", "cash_vehicle_subsidy",
        "cash_self_development", "cash_homecoming_travel", "cash_welfare_points",
        "cash_flexible_welfare_points", "cash_fitness", "cash_pension",
    ])
    offered_total = offered_contract + data["offer_incentive"] + cash_benefits

    def compare(offered: float, current: float) -> tuple[float, float, str]:
        diff = offered - current
        rate = diff / current * 100 if current else 0
        return diff, rate, "상향" if diff > 0 else "하향" if diff < 0 else "동일"

    contract_difference, contract_rate, contract_result = compare(offered_contract, data["current_contract_salary"])
    total_difference, total_rate, total_result = compare(offered_total, data["current_total_salary"])
    grade_group = pay_grade_group(st.session_state.c_grade, st.session_state.c_grade_role)
    band_row = get_pay_band_row(grade_group, st.session_state.c_band) if st.session_state.c_band else None

    return {
        **data,
        "grade_group": grade_group,
        "band_row": band_row,
        "offered_contract": offered_contract,
        "cash_benefits": cash_benefits,
        "offered_total": offered_total,
        "first_year_total": offered_total + data["sign_on_bonus"],
        "contract_difference": contract_difference,
        "contract_difference_rate": contract_rate,
        "contract_result": contract_result,
        "difference": total_difference,
        "difference_rate": total_rate,
        "result": total_result,
    }


def format_period_short(total_days: float | int) -> str:
    years, months, days = days_to_ymd(total_days)
    parts = []
    if years:
        parts.append(f"{years}년")
    if months:
        parts.append(f"{months}개월")
    if not parts:
        parts.append(f"{days}일")
    return " ".join(parts)


def image_to_data_uri(path_text: Any) -> str:
    path = Path(str(path_text or ""))
    if not path.exists() or not path.is_file():
        return ""
    mime = mimetypes.guess_type(path.name)[0] or "image/png"
    try:
        encoded = base64.b64encode(path.read_bytes()).decode("ascii")
        return f"data:{mime};base64,{encoded}"
    except Exception:
        return ""


def candidate_age_text(birth_date: Any, 기준일: date | None = None) -> str:
    parsed = pd.to_datetime(birth_date, errors="coerce")
    if pd.isna(parsed):
        return ""
    today = 기준일 or date.today()
    b = parsed.date()
    age = today.year - b.year - ((today.month, today.day) < (b.month, b.day))
    return f"만 {age}세"


def career_line_summary(career_df: pd.DataFrame) -> str:
    df = ensure_career_columns(career_df)
    if df.empty:
        return ""
    df["입사일"] = pd.to_datetime(df["입사일"], errors="coerce")
    df["퇴사일"] = pd.to_datetime(df["퇴사일"], errors="coerce")
    rows = []
    for _, row in df.sort_values(["입사일", "회사명"]).iterrows():
        company = str(row.get("표준회사명") or row.get("회사명") or "").strip()
        if not company or pd.isna(row.get("입사일")):
            continue
        start = row["입사일"].date()
        end_val = row.get("퇴사일")
        end = end_val.date() if not pd.isna(end_val) else date.today()
        rate = float(pd.to_numeric(row.get("인정률(%)", 0), errors="coerce") or 0)
        days = max((end - start).days + 1, 0) * rate / 100
        rows.append(
            f"{company}(`{start:%y.%m}~`{end:%y.%m}) /{format_period_short(days)} ({rate:.0f}%)"
        )
    return "\n".join(rows)


def compensation_one_line(compensation: dict[str, Any] | None) -> str:
    comp = compensation or {}
    grade = str(comp.get("selected_grade") or "").strip()
    selected_band = str(comp.get("selected_band") or "").strip()
    band = selected_band.split("|", 1)[1] if "|" in selected_band else selected_band
    contract = sum(float(comp.get(k, 0) or 0) for k in ["offer_base_salary", "offer_performance_salary", "offer_fixed_overtime"])
    promo = str(comp.get("promotion_base_date") or "").strip()
    head = f"{grade}{f'({band})' if band else ''} / {format_won(contract)}" if grade else format_won(contract)
    return head + (f"\n※ 승급기준일 : `{promo}`" if promo else "")


def dashboard_rows(candidates: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for candidate in candidates:
        candidate_id = int(candidate["id"])
        career_df = ensure_career_columns(pd.DataFrame(get_career_records(candidate_id)))
        career_result = calculate_career(career_df, date.today())
        compensation = get_compensation(candidate_id) or {}
        rows.append({
            "사진": image_to_data_uri(candidate.get("photo_path", "")),
            "성명": candidate.get("name", ""),
            "성별": candidate.get("gender", ""),
            "연령": candidate_age_text(candidate.get("birth_date", "")),
            "학력": candidate.get("education", ""),
            "전공": candidate.get("major", ""),
            "경력사항": career_line_summary(career_df),
            "인정 기간": format_period_short(career_result["recognized_days"]),
            "처우": compensation_one_line(compensation),
            "후보자번호": candidate.get("candidate_code", ""),
            "진행상태": candidate.get("status", ""),
        })
    return pd.DataFrame(rows, columns=["사진", "성명", "성별", "연령", "학력", "전공", "경력사항", "인정 기간", "처우", "후보자번호", "진행상태"])


def dashboard_excel_bytes(df: pd.DataFrame, candidates: list[dict[str, Any]]) -> bytes:
    # 사진은 가능하면 실제 이미지로 삽입하고, 실패하면 텍스트 표만 제공합니다.
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "처우계산 현황"
    headers = ["사진", "성명", "성별", "연령", "학력", "전공", "경력사항", "인정 기간", "처우"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="D9E2F3")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row_idx, row in enumerate(df.to_dict("records"), 2):
        for col_idx, header in enumerate(headers, 1):
            if header == "사진":
                ws.cell(row_idx, col_idx).value = ""
            else:
                ws.cell(row_idx, col_idx).value = row.get(header, "")
            ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row_idx, col_idx).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        # 후보자 원본의 photo_path로 이미지 삽입
        try:
            original = candidates[row_idx - 2]
            photo_path = Path(str(original.get("photo_path") or ""))
            if photo_path.exists():
                img = XLImage(str(photo_path))
                img.width = 52
                img.height = 64
                ws.add_image(img, f"A{row_idx}")
        except Exception:
            pass
        ws.row_dimensions[row_idx].height = 58
    widths = [10, 12, 8, 10, 14, 16, 48, 14, 28]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


init_db()
st.set_page_config(page_title=APP_TITLE, page_icon="📋", layout="wide")
ensure_session_defaults()

st.markdown(
    """
    <style>
      .block-container {padding-top: 1.4rem; padding-bottom: 3rem;}
      .small-note {font-size: 0.85rem; color: #666;}
      .info-box {
        border: 1px solid #d9e2f3; border-radius: 8px;
        padding: 14px; background: #f7f9fc;
      }
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.title("처우검토 시스템")
    nav = st.radio(
        "메뉴",
        ["후보자 목록", "후보자 검토", "관리자 설정", "시스템 안내"],
        key="nav_page",
    )
    st.divider()
    st.button(
        "➕ 신규 후보자",
        use_container_width=True,
        on_click=go_to_new_candidate,
    )

    if st.session_state.current_candidate_id:
        st.caption(f"현재 후보자: {st.session_state.candidate_code}")
        st.button(
            "현재 후보자 다시 불러오기",
            use_container_width=True,
            on_click=reload_current_candidate,
        )


if nav == "후보자 목록":
    st.title("처우계산 대상자 현황")
    st.caption("현재 처우계산 대상자를 한눈에 보고, 현황표를 Excel로 다운로드합니다.")

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        search_text = st.text_input(
            "검색",
            placeholder="성명, 후보자번호, 현 직장, 지원직무",
        )
    with col2:
        status_filter = st.selectbox("상태", ["전체"] + CANDIDATE_STATUSES)
    with col3:
        st.write("")
        st.write("")
        st.button(
            "새 후보자 등록",
            use_container_width=True,
            on_click=go_to_new_candidate,
        )

    candidates = list_candidates(search_text, status_filter)
    if not candidates:
        st.info("저장된 처우계산 대상자가 없습니다.")
    else:
        dashboard_df = dashboard_rows(candidates)
        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        kpi1.metric("전체 대상자", f"{len(dashboard_df):,}명")
        kpi2.metric("검토중", f"{sum(dashboard_df['진행상태'] == '검토중'):,}명")
        kpi3.metric("처우협의중", f"{sum(dashboard_df['진행상태'] == '처우협의중'):,}명")
        kpi4.metric("오퍼발송/확정", f"{sum(dashboard_df['진행상태'].isin(['오퍼발송', '입사확정'])):,}명")

        view_df = dashboard_df[["사진", "성명", "성별", "연령", "학력", "전공", "경력사항", "인정 기간", "처우"]]
        st.dataframe(
            view_df,
            hide_index=True,
            use_container_width=True,
            height=520,
            column_config={
                "사진": st.column_config.ImageColumn("사진", width="small"),
                "성명": st.column_config.TextColumn("성명", width="small"),
                "성별": st.column_config.TextColumn("성별", width="small"),
                "연령": st.column_config.TextColumn("연령", width="small"),
                "학력": st.column_config.TextColumn("학력", width="small"),
                "전공": st.column_config.TextColumn("전공", width="medium"),
                "경력사항": st.column_config.TextColumn("경력사항", width="large"),
                "인정 기간": st.column_config.TextColumn("인정 기간", width="small"),
                "처우": st.column_config.TextColumn("처우", width="medium"),
            },
        )

        st.download_button(
            "처우계산 대상자 현황 Excel 다운로드",
            data=dashboard_excel_bytes(dashboard_df, candidates),
            file_name=f"처우계산_대상자_현황_{date.today():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        option_map = {
            f"{row['candidate_code']} | {row['name']} | {row.get('target_job', '')}": row["id"]
            for row in candidates
        }
        selected_label = st.selectbox("상세조회할 후보자", list(option_map.keys()))
        st.button(
            "선택 후보자 열기",
            type="primary",
            on_click=open_candidate_and_go,
            args=(int(option_map[selected_label]),),
        )

elif nav == "후보자 검토":
    title_name = st.session_state.f_name or "신규 후보자"
    st.title(f"후보자 검토 · {title_name}")
    if st.session_state.candidate_code:
        st.caption(f"후보자번호: {st.session_state.candidate_code}")
    else:
        st.caption("아직 저장되지 않은 신규 후보자입니다.")

    top_col1, top_col2 = st.columns([1, 5])
    with top_col1:
        if st.button("💾 전체 저장", type="primary", use_container_width=True):
            save_current_candidate()
    with top_col2:
        st.caption(
            "전체 저장 시 기본정보·경력·보상·복리후생이 DB에 저장되고 스냅샷과 변경이력이 남습니다."
        )

    tabs = st.tabs([
        "① 기본정보",
        "② 증빙문서·OCR",
        "③ 인정경력",
        "④ 보상비교",
        "⑤ 복리후생",
        "⑥ 오퍼레터",
        "⑦ 변경이력",
    ])

    with tabs[0]:
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.text_input("후보자 성명 *", key="f_name")
            st.selectbox("성별", ["", "남", "여"], key="f_gender")
            birth_value = st.session_state.f_birth_date or date(1990, 1, 1)
            birth_checked = st.checkbox("생년월일 입력", value=st.session_state.f_birth_date is not None)
            st.session_state.f_birth_date = st.date_input("생년월일", value=birth_value, disabled=not birth_checked) if birth_checked else None
        with col2:
            st.text_input("이메일", key="f_email")
            st.text_input("연락처", key="f_phone")
            st.text_input("학력", key="f_education", placeholder="예: 학사 / 석사 / 박사")
            st.text_input("전공", key="f_major")
        with col3:
            st.text_input("현 직장", key="f_current_company")
            st.text_input("채용 부서", key="f_department")
            st.text_input("지원 직무/직위", key="f_target_job")
            st.selectbox("직군", JOB_GROUP_OPTIONS, key="f_job_group")
        with col4:
            st.selectbox("계약구분", EMPLOYMENT_TYPES, key="f_employment_type")
            st.selectbox("근무지", WORK_LOCATION_OPTIONS, key="f_work_location")
            st.selectbox("입사구분", ENTRY_TYPE_OPTIONS, key="f_entry_type")
            st.date_input("입사 예정일", key="f_expected_join_date")
            st.caption("직급은 ④ 보상비교의 Pay Band 직급을 복리후생 기준에도 사용합니다.")

        photo_col, note_col = st.columns([1, 3])
        with photo_col:
            uploaded_photo = st.file_uploader("사진 업로드", type=["png", "jpg", "jpeg"], key="candidate_photo_upload")
            if uploaded_photo is not None:
                photo_dir = UPLOAD_DIR / "candidate_photos"
                photo_dir.mkdir(parents=True, exist_ok=True)
                photo_name = f"{safe_filename(st.session_state.candidate_code or st.session_state.f_name or 'candidate')}_{datetime.now():%Y%m%d_%H%M%S}_{safe_filename(uploaded_photo.name)}"
                photo_path = photo_dir / photo_name
                photo_path.write_bytes(uploaded_photo.getvalue())
                st.session_state.f_photo_path = str(photo_path)
                st.success("사진을 저장했습니다. 전체 저장을 눌러 후보자 기록에 반영하세요.")
            if st.session_state.f_photo_path and Path(st.session_state.f_photo_path).exists():
                st.image(st.session_state.f_photo_path, width=120)
        with note_col:
            c1, c2 = st.columns([1, 3])
            with c1:
                st.selectbox("진행 상태", CANDIDATE_STATUSES, key="f_status")
            with c2:
                st.text_area("검토 메모", key="f_notes", height=110)

    with tabs[1]:
        st.subheader("건강보험 자격득실확인서 및 증빙문서")
        st.info(
            "PDF 내부 문자추출 결과의 품질을 먼저 검사하고, 날짜·한글·건강보험 핵심항목이 충분하지 않으면 "
            "자동으로 320~400dpi OCR, 대비향상, 이진화, 기울기/회전 보정, 복수 PSM을 재시도합니다. "
            "자동 추출 결과는 반드시 원본과 대조한 뒤 인정률을 확정하세요."
        )

        status = tesseract_status()
        if status["available"]:
            if status.get("has_korean", False):
                st.success(f"OCR 사용 가능 · {status.get('version', '')}")
            else:
                st.warning(status["message"])
        else:
            st.warning(status["message"])

        if not st.session_state.current_candidate_id:
            st.warning("문서를 올리기 전에 기본정보에서 후보자 성명을 입력하고 전체 저장을 먼저 하세요.")
        else:
            document_type = st.selectbox(
                "문서 유형",
                ["건강보험 자격득실확인서", "경력증명서", "재직증명서", "연봉자료", "기타"],
            )
            uploaded_files = st.file_uploader(
                "PDF 또는 이미지 업로드",
                type=["pdf", "png", "jpg", "jpeg", "tif", "tiff", "bmp"],
                accept_multiple_files=True,
            )
            if st.button("업로드 파일 저장", disabled=not uploaded_files):
                for uploaded in uploaded_files or []:
                    target = save_uploaded_file(
                        st.session_state.candidate_code,
                        uploaded.name,
                        uploaded.getvalue(),
                    )
                    add_document(
                        st.session_state.current_candidate_id,
                        document_type,
                        uploaded.name,
                        str(target),
                    )
                    add_audit_log(
                        st.session_state.current_candidate_id,
                        "문서 업로드",
                        {"document_type": document_type, "file_name": uploaded.name},
                    )
                st.success(f"{len(uploaded_files or [])}개 파일을 저장했습니다.")
                st.rerun()

            documents = get_documents(st.session_state.current_candidate_id)
            if not documents:
                st.caption("저장된 문서가 없습니다.")
            else:
                document_df = pd.DataFrame(documents)
                document_df["OCR"] = document_df["ocr_used"].map({0: "미사용", 1: "사용"})
                st.dataframe(
                    document_df[
                        [
                            "id", "document_type", "original_name", "OCR",
                            "uploaded_at", "analyzed_at", "analysis_message",
                        ]
                    ].rename(columns={
                        "id": "문서ID",
                        "document_type": "유형",
                        "original_name": "파일명",
                        "uploaded_at": "업로드일시",
                        "analyzed_at": "분석일시",
                        "analysis_message": "분석결과",
                    }),
                    hide_index=True,
                    use_container_width=True,
                )

                doc_option_map = {
                    f"{doc['id']} | {doc['document_type']} | {doc['original_name']}": doc["id"]
                    for doc in documents
                }
                selected_doc_label = st.selectbox(
                    "분석 또는 확인할 문서",
                    list(doc_option_map.keys()),
                )
                selected_document = get_document(int(doc_option_map[selected_doc_label]))

                option1, option2 = st.columns(2)
                with option1:
                    selected_suffix = (
                        Path(selected_document["stored_path"]).suffix.lower()
                        if selected_document
                        else ""
                    )
                    pdf_password = ""
                    if selected_suffix == ".pdf":
                        pdf_password = st.text_input(
                            "PDF 비밀번호 — 생년월일 6자리",
                            type="password",
                            max_chars=6,
                            placeholder="예: 900101",
                            help=(
                                "암호화된 건강보험 자격득실확인서는 생년월일 6자리를 입력하세요. "
                                "입력값은 DB나 변경이력에 저장하지 않습니다."
                            ),
                            key=f"pdf_password_{selected_document['id']}",
                        )
                    force_ocr = st.checkbox(
                        "고급 OCR을 무조건 다시 실행",
                        value=False,
                        help=(
                            "기본 상태에서도 직접추출 품질이 낮으면 OCR을 자동 실행합니다. "
                            "원문이 깨져 보이거나 자동 추출 행이 누락될 때만 체크하세요."
                        ),
                    )
                    subtract_one_day = st.checkbox(
                        "자격상실일에서 1일 차감",
                        value=True,
                        help="건강보험 자격상실일이 실제 퇴사일 다음 날인 경우를 반영합니다.",
                    )
                    application_deadline = st.date_input(
                        "지원서 마감일 (재직중 경력 종료일)",
                        value=date.today(),
                        help=(
                            "건강보험 자격상실일이 비어 있는 직장가입자는 재직중으로 판단하고, "
                            "이 날짜까지를 경력으로 계산합니다. 기존 회사의 상실일에는 적용하지 않습니다."
                        ),
                        key=f"application_deadline_{selected_document['id']}",
                    )
                with option2:
                    st.caption(
                        "암호가 없는 PDF는 비밀번호 입력란을 비워두면 됩니다. "
                        "암호화 PDF는 분석 시 입력한 값으로만 일시적으로 인증합니다."
                    )
                    if selected_document:
                        path = Path(selected_document["stored_path"])
                        if path.exists():
                            st.download_button(
                                "원본 문서 다운로드",
                                data=path.read_bytes(),
                                file_name=selected_document["original_name"],
                                mime="application/octet-stream",
                                use_container_width=True,
                            )

                if st.button("선택 문서 분석 및 경력표 갱신", type="primary"):
                    if not selected_document:
                        st.error("문서를 선택하세요.")
                    else:
                        try:
                            with st.spinner("문서에서 문자를 추출하고 경력 행을 분석하고 있습니다."):
                                extract_kwargs = {"force_ocr": force_ocr}
                                if OCR_PASSWORD_SUPPORTED:
                                    extract_kwargs["password"] = pdf_password
                                elif pdf_password:
                                    raise RuntimeError(
                                        "ocr_service.py가 이전 버전입니다. "
                                        "프로젝트 폴더의 ocr_service.py를 v2.2.1 파일로 교체하세요."
                                    )

                                is_health_document = (
                                    selected_document.get("document_type") == "건강보험 자격득실확인서"
                                )
                                if OCR_HEALTH_HINT_SUPPORTED:
                                    extract_kwargs["health_document"] = is_health_document

                                result = extract_text_from_file(
                                    selected_document["stored_path"],
                                    **extract_kwargs,
                                )
                                parsed_df = parse_health_insurance_career(
                                    result["text"],
                                    subtract_one_day_from_loss_date=subtract_one_day,
                                    current_employment_end_date=application_deadline,
                                ) if is_health_document else pd.DataFrame()
                                parse_diag = health_parse_diagnostics(result["text"], parsed_df)

                                # 직접추출이 그럴듯해 보여도 실제 경력행 파싱률이 낮으면
                                # 고급 OCR을 한 번 더 강제 실행하고 더 좋은 결과를 자동 선택합니다.
                                needs_auto_retry = (
                                    is_health_document
                                    and status.get("available", False)
                                    and not force_ocr
                                    and (
                                        parsed_df.empty
                                        or parse_diag.get("score", 0) < 48
                                        or (
                                            parse_diag.get("expected_pairs_from_dates", 0) >= 2
                                            and parse_diag.get("parse_coverage", 0) < 0.60
                                        )
                                    )
                                )
                                if needs_auto_retry:
                                    retry_kwargs = dict(extract_kwargs)
                                    retry_kwargs["force_ocr"] = True
                                    retry_result = extract_text_from_file(
                                        selected_document["stored_path"],
                                        **retry_kwargs,
                                    )
                                    retry_df = parse_health_insurance_career(
                                        retry_result["text"],
                                        subtract_one_day_from_loss_date=subtract_one_day,
                                        current_employment_end_date=application_deadline,
                                    )
                                    retry_diag = health_parse_diagnostics(retry_result["text"], retry_df)

                                    # 경력행 수를 우선하고, 같으면 품질점수가 높은 결과를 채택합니다.
                                    if (
                                        len(retry_df) > len(parsed_df)
                                        or (
                                            len(retry_df) == len(parsed_df)
                                            and retry_diag.get("score", 0) > parse_diag.get("score", 0)
                                        )
                                    ):
                                        result = retry_result
                                        parsed_df = retry_df
                                        parse_diag = retry_diag
                                        result["message"] += " · 경력행 검증 실패로 고급 OCR 자동 재시도 결과를 채택"

                                if not parsed_df.empty:
                                    parsed_df["출처파일"] = selected_document["original_name"]

                                page_details = result.get("diagnostics", {}).get("pages", [])
                                page_summary = ", ".join(
                                    f"P{item.get('page')} {item.get('method')}({item.get('score', 0)})"
                                    for item in page_details[:6]
                                )
                                analysis_message = (
                                    f"{result['message']} · 경력후보 {len(parsed_df)}건"
                                    + (f" · {page_summary}" if page_summary else "")
                                )

                                update_document_analysis(
                                    selected_document["id"],
                                    result["text"],
                                    bool(result["ocr_used"]),
                                    analysis_message,
                                )

                                if not parsed_df.empty:
                                    current_df = ensure_career_columns(st.session_state.career_df.copy())

                                    # 건강보험 자격득실확인서는 같은 문서를 재분석할 때
                                    # 과거 OCR 결과를 누적하지 않고 새 결과로 교체합니다.
                                    removed_rows = 0
                                    if is_health_document and not current_df.empty:
                                        source_col = current_df["출처"].fillna("").astype(str).str.strip()
                                        file_col = current_df["출처파일"].fillna("").astype(str).str.strip()
                                        selected_name = str(selected_document["original_name"]).strip()

                                        same_health_source = source_col.eq("건강보험 OCR")
                                        same_file = file_col.eq(selected_name)
                                        legacy_unassigned = same_health_source & file_col.isin(["", "nan", "<NA>", "None"])

                                        # 동일 파일의 예전 OCR 결과 + 출처파일이 없던 구버전 OCR 결과는 제거.
                                        replace_mask = same_health_source & (same_file | legacy_unassigned)
                                        removed_rows = int(replace_mask.sum())
                                        current_df = current_df.loc[~replace_mask].copy()

                                    parsed_df = ensure_career_columns(parsed_df)
                                    combined = pd.concat([current_df, parsed_df], ignore_index=True)

                                    # 같은 회사/입사일/퇴사일/출처파일이 완전히 같은 행은 1건만 유지합니다.
                                    dedupe_cols = ["회사명", "입사일", "퇴사일", "출처", "출처파일"]
                                    for col in dedupe_cols:
                                        if col not in combined.columns:
                                            combined[col] = ""
                                    combined = combined.drop_duplicates(subset=dedupe_cols, keep="last").reset_index(drop=True)

                                    st.session_state.career_df = ensure_career_columns(combined)
                                    st.session_state.career_editor_version += 1

                                    # OCR 재분석 결과를 즉시 DB에도 반영해 구버전의 잘못된 행이
                                    # 다음 화면 진입 때 다시 살아나지 않도록 합니다.
                                    if st.session_state.current_candidate_id:
                                        replace_career_records(
                                            st.session_state.current_candidate_id,
                                            career_records_for_db(st.session_state.career_df),
                                        )

                                    add_audit_log(
                                        st.session_state.current_candidate_id,
                                        "OCR 경력 추출",
                                        {
                                            "file_name": selected_document["original_name"],
                                            "extracted_rows": len(parsed_df),
                                            "removed_previous_ocr_rows": removed_rows,
                                            "ocr_used": result["ocr_used"],
                                        },
                                    )
                                    replace_note = f"기존 OCR {removed_rows}건 제거 후 " if removed_rows else ""
                                    st.success(
                                        f"{replace_note}{len(parsed_df)}개 경력 후보로 갱신했습니다. "
                                        "인정경력 탭에서 반드시 확인하세요."
                                    )
                                else:
                                    reason = []
                                    if not status.get("available", False):
                                        reason.append("Tesseract 미설치")
                                    elif not status.get("has_korean", False):
                                        reason.append("한국어(kor) 언어데이터 없음")
                                    if parse_diag.get("date_count", 0) < 2:
                                        reason.append("취득/상실 날짜 패턴 인식 부족")
                                    if parse_diag.get("hangul_count", 0) < 4:
                                        reason.append("한글 회사명 인식 부족")
                                    reason_text = ", ".join(reason) if reason else "문서 표 구조 또는 OCR 품질 문제"
                                    st.warning(
                                        "자동으로 경력행을 찾지 못했습니다. "
                                        f"원인 후보: {reason_text}. 저장된 추출문자를 확인하거나 고급 OCR을 체크해 재실행하세요."
                                    )
                                st.rerun()
                        except Exception as exc:
                            st.error(f"문서 분석 중 오류: {exc}")

                if selected_document and selected_document.get("extracted_text"):
                    with st.expander("저장된 추출문자 보기"):
                        st.text_area(
                            "추출문자",
                            value=selected_document["extracted_text"],
                            height=350,
                            disabled=True,
                        )

    with tabs[2]:
        st.subheader("인정경력 산정")
        col1, col2 = st.columns([1, 2])
        with col1:
            calculation_date = st.date_input("경력 산정기준일", value=date.today())
        with col2:
            st.caption(
                "직전년도 연결매출을 DART에서 조회하고 관리자 매출기준을 적용합니다. "
                "회사 매칭·연결매출·기준구간 중 하나라도 확인되지 않으면 확인필요로 남고 인정경력에서 제외됩니다."
            )

        action1, action2 = st.columns([1, 2])
        with action1:
            run_dart = st.button(
                "DART 연결매출 조회·인정률 일괄반영",
                type="primary",
                use_container_width=True,
            )
        with action2:
            st.caption(
                f"조회대상 사업연도: {date.today().year - 1}년 · 연결재무제표(CFS) · 사업보고서"
            )

        if run_dart:
            api_key = get_setting("DART_API_KEY", "").strip()
            if not api_key:
                st.error("관리자 설정에서 DART API 키를 먼저 저장하세요.")
            elif st.session_state.career_df.empty:
                st.warning("조회할 경력 회사가 없습니다.")
            else:
                try:
                    with st.spinner("DART 회사 매칭과 연결매출 조회를 진행하고 있습니다."):
                        if not DART_CORP_CACHE.exists():
                            download_corp_codes(api_key, DART_CORP_CACHE)
                        corp_df = load_corp_codes(DART_CORP_CACHE)
                        rules = get_revenue_rules()
                        if not rules:
                            raise RuntimeError("관리자 매출 인정기준이 없습니다.")

                        result_df = ensure_career_columns(st.session_state.career_df)
                        target_year = date.today().year - 1
                        completed = 0
                        pending = 0

                        for idx, row in result_df.iterrows():
                            company_name = str(row.get("회사명", "") or "").strip()
                            if not company_name:
                                continue
                            if bool(row.get("수동확정", False)):
                                result_df.at[idx, "인정구분"] = "수동확정"
                                if str(row.get("외부자료URL", "") or "").strip() or str(row.get("외부자료출처", "") or "").strip():
                                    result_df.at[idx, "매출조회상태"] = "외부자료확정"
                                else:
                                    result_df.at[idx, "매출조회상태"] = "수동확정"
                                completed += 1
                                continue

                            saved = get_company_mapping(company_name) or {}
                            manual_code = normalize_corp_code(row.get("DART고유번호", ""))
                            match = find_company_match(
                                company_name,
                                corp_df,
                                mapped_corp_code=manual_code or str(saved.get("corp_code", "")),
                                mapped_corp_name=str(saved.get("corp_name", "")),
                            )
                            if not match.get("matched"):
                                result_df.at[idx, "표준회사명"] = ""
                                result_df.at[idx, "매출연도"] = target_year
                                result_df.at[idx, "연결매출(억원)"] = pd.NA
                                result_df.at[idx, "매출기준인정률(%)"] = pd.NA
                                result_df.at[idx, "매출조회상태"] = match.get("status", "확인필요")
                                result_df.at[idx, "인정구분"] = "확인필요"
                                result_df.at[idx, "인정률(%)"] = 0
                                suggestions = match.get("suggestions") or []
                                if suggestions:
                                    suggestion_text = ", ".join(
                                        f"{item.get('corp_name')}({item.get('corp_code')})"
                                        for item in suggestions[:3]
                                    )
                                    result_df.at[idx, "비고"] = (
                                        str(row.get("비고", "") or "")
                                        + f" | DART 후보: {suggestion_text}"
                                    ).strip(" |")
                                pending += 1
                                continue

                            corp_code = str(match["corp_code"])
                            corp_name = str(match["corp_name"])
                            result_df.at[idx, "표준회사명"] = corp_name
                            result_df.at[idx, "DART고유번호"] = corp_code
                            result_df.at[idx, "매출연도"] = target_year
                            upsert_company_mapping(
                                company_name,
                                normalize_company_name(company_name),
                                corp_code,
                                corp_name,
                            )

                            revenue = fetch_consolidated_revenue(
                                api_key, corp_code, target_year,
                            )
                            if not revenue.get("ok"):
                                result_df.at[idx, "연결매출(억원)"] = pd.NA
                                result_df.at[idx, "매출기준인정률(%)"] = pd.NA
                                result_df.at[idx, "매출조회상태"] = revenue.get("status", "확인필요")
                                result_df.at[idx, "인정구분"] = "확인필요"
                                result_df.at[idx, "인정률(%)"] = 0
                                pending += 1
                                continue

                            rule = apply_revenue_rule(float(revenue["revenue_eok"]), rules)
                            if not rule:
                                result_df.at[idx, "연결매출(억원)"] = float(revenue["revenue_eok"])
                                result_df.at[idx, "매출조회상태"] = "확인필요: 매출 기준구간 없음"
                                result_df.at[idx, "인정구분"] = "확인필요"
                                result_df.at[idx, "인정률(%)"] = 0
                                pending += 1
                                continue

                            rate = float(rule["recognition_rate"])
                            result_df.at[idx, "연결매출(억원)"] = round(float(revenue["revenue_eok"]), 2)
                            result_df.at[idx, "매출기준인정률(%)"] = rate
                            result_df.at[idx, "공시접수번호"] = str(revenue.get("rcept_no", ""))
                            result_df.at[idx, "매출조회상태"] = "조회완료"
                            result_df.at[idx, "인정구분"] = "매출기준"
                            result_df.at[idx, "인정률(%)"] = rate
                            completed += 1

                        st.session_state.career_df = result_df
                        st.session_state.career_editor_version += 1
                        if st.session_state.current_candidate_id:
                            add_audit_log(
                                st.session_state.current_candidate_id,
                                "DART 연결매출 및 인정률 일괄반영",
                                {"사업연도": target_year, "완료": completed, "확인필요": pending},
                            )
                        st.success(f"조회완료 {completed}건, 확인필요 {pending}건")
                        st.rerun()
                except Exception as exc:
                    st.error(f"DART 일괄조회 중 오류: {exc}")

        edited_career_df = st.data_editor(
            ensure_career_columns(st.session_state.career_df),
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            key=f"career_editor_{st.session_state.career_editor_version}",
            column_config={
                "회사명": st.column_config.TextColumn("건강보험상 회사명", required=True),
                "입사일": st.column_config.DateColumn("입사일", format="YYYY-MM-DD", required=True),
                "퇴사일": st.column_config.DateColumn("퇴사일", format="YYYY-MM-DD"),
                "직무/직책": st.column_config.TextColumn("직무/직책"),
                "표준회사명": st.column_config.TextColumn("DART 회사명"),
                "DART고유번호": st.column_config.TextColumn(
                    "DART 고유번호",
                    help="자동매칭 실패 시 관리자 DART 기업목록에서 확인한 8자리 번호를 입력하고 다시 조회할 수 있습니다.",
                ),
                "매출연도": st.column_config.NumberColumn("매출연도", format="%d"),
                "연결매출(억원)": st.column_config.NumberColumn("연결매출(억원)", format="%.0f"),
                "매출조회상태": st.column_config.TextColumn("매출조회상태"),
                "매출기준인정률(%)": st.column_config.NumberColumn("자동 인정률", format="%.0f%%"),
                "공시접수번호": st.column_config.TextColumn("공시접수번호"),
                "외부자료출처": st.column_config.TextColumn("외부자료 출처"),
                "외부자료URL": st.column_config.LinkColumn("외부자료 URL", display_text="링크"),
                "재무기준": st.column_config.TextColumn("재무기준"),
                "외부증빙파일": st.column_config.TextColumn("외부 증빙파일"),
                "수동확정": st.column_config.CheckboxColumn(
                    "수동확정",
                    help="DART로 확인할 수 없는 경우에만 체크하고 인정률과 조정사유를 직접 입력합니다.",
                ),
                "수동조정사유": st.column_config.TextColumn("수동확정 사유"),
                "인정구분": st.column_config.SelectboxColumn(
                    "인정구분", options=list(CAREER_RATE_GUIDE.keys()), required=True,
                ),
                "인정률(%)": st.column_config.NumberColumn(
                    "최종 인정률", min_value=0, max_value=100, step=5, format="%d%%",
                ),
                "출처": st.column_config.SelectboxColumn(
                    "출처", options=["수기입력", "건강보험 OCR", "경력증명서", "외부자료 수동확정", "기타"],
                ),
                "출처파일": st.column_config.TextColumn("출처파일"),
                "비고": st.column_config.TextColumn("비고"),
            },
            disabled=[
                "표준회사명", "매출연도", "연결매출(억원)",
                "매출조회상태", "매출기준인정률(%)", "공시접수번호",
                "외부자료출처", "외부자료URL", "재무기준", "외부증빙파일",
            ],
        )
        st.session_state.career_df = ensure_career_columns(edited_career_df)

        st.markdown("#### DART 미확인 회사 · 외부자료 확정")
        st.caption(
            "DART에서 연결매출을 확인하지 못한 회사는 검색 결과나 회사 공식자료를 확인한 뒤 "
            "직전년도 연결매출을 입력해 확정할 수 있습니다. 입력한 매출액에는 관리자 매출 인정기준이 자동 적용됩니다."
        )
        career_for_external = ensure_career_columns(st.session_state.career_df)
        candidate_rows = []
        for ext_idx, ext_row in career_for_external.iterrows():
            ext_company = str(ext_row.get("회사명", "") or "").strip()
            if not ext_company:
                continue
            ext_status = str(ext_row.get("매출조회상태", "") or "미조회")
            if ext_status == "조회완료" and not bool(ext_row.get("수동확정", False)):
                continue
            candidate_rows.append((ext_idx, ext_company, ext_status))

        if candidate_rows:
            option_map = {
                f"{int(idx) + 1}. {company} · {status}": idx
                for idx, company, status in candidate_rows
            }
            selected_external_label = st.selectbox(
                "외부자료로 확인할 경력회사",
                list(option_map.keys()),
                key="external_revenue_target",
            )
            selected_external_idx = option_map[selected_external_label]
            external_row = career_for_external.loc[selected_external_idx]
            external_company = str(external_row.get("회사명", "") or "").strip()
            default_year = int(pd.to_numeric(external_row.get("매출연도"), errors="coerce")) if pd.notna(pd.to_numeric(external_row.get("매출연도"), errors="coerce")) else date.today().year - 1

            search_query = quote_plus(f'{external_company} {default_year} 연결 매출액')
            search_col1, search_col2, search_col3 = st.columns([1, 1, 2])
            with search_col1:
                st.link_button(
                    "Google 검색",
                    f"https://www.google.com/search?q={search_query}",
                    use_container_width=True,
                )
            with search_col2:
                st.link_button(
                    "Naver 검색",
                    f"https://search.naver.com/search.naver?query={search_query}",
                    use_container_width=True,
                )
            with search_col3:
                st.caption(f"검색어: {external_company} {default_year} 연결 매출액")

            ext_col1, ext_col2, ext_col3 = st.columns(3)
            with ext_col1:
                external_source_type = st.selectbox(
                    "외부자료 출처",
                    ["회사 공식 홈페이지", "감사보고서/공시자료", "기업정보서비스", "언론기사", "검색결과 기타"],
                    key=f"external_source_type_{selected_external_idx}",
                )
                external_year = st.number_input(
                    "매출 기준연도",
                    min_value=2000,
                    max_value=date.today().year,
                    value=default_year,
                    step=1,
                    key=f"external_year_{selected_external_idx}",
                )
            with ext_col2:
                external_statement_type = st.selectbox(
                    "재무기준",
                    ["연결", "별도", "구분확인필요"],
                    help="현재 경력 인정기준이 연결매출 기준이므로 '연결'이 확인된 자료만 확정됩니다.",
                    key=f"external_statement_{selected_external_idx}",
                )
                previous_revenue = pd.to_numeric(external_row.get("연결매출(억원)"), errors="coerce")
                external_revenue = st.number_input(
                    "매출액(억원)",
                    min_value=0.0,
                    value=float(previous_revenue) if pd.notna(previous_revenue) else 0.0,
                    step=100.0,
                    format="%.1f",
                    key=f"external_revenue_{selected_external_idx}",
                )
            with ext_col3:
                external_url = st.text_input(
                    "근거 URL",
                    value=str(external_row.get("외부자료URL", "") or ""),
                    placeholder="https://...",
                    key=f"external_url_{selected_external_idx}",
                )
                external_reason = st.text_input(
                    "확인 메모",
                    value=str(external_row.get("수동조정사유", "") or ""),
                    placeholder="예: 2025년 연결 매출액 확인",
                    key=f"external_reason_{selected_external_idx}",
                )

            external_evidence = st.file_uploader(
                "외부자료 증빙파일 첨부(선택)",
                type=["pdf", "png", "jpg", "jpeg", "xlsx", "xls", "docx"],
                key=f"external_evidence_{selected_external_idx}",
            )

            if st.button("외부자료 매출 확정 · 인정률 자동적용", type="primary"):
                ext_errors = []
                if external_statement_type != "연결":
                    ext_errors.append("연결 기준 매출인지 확인된 자료만 확정할 수 있습니다.")
                if external_revenue <= 0:
                    ext_errors.append("매출액을 0보다 크게 입력하세요.")
                if not external_url.strip() and external_evidence is None:
                    ext_errors.append("근거 URL 또는 증빙파일 중 하나는 입력하세요.")

                rules = get_revenue_rules()
                ext_rule = apply_revenue_rule(float(external_revenue), rules) if not ext_errors else None
                if not ext_errors and not ext_rule:
                    ext_errors.append("현재 관리자 매출 인정기준에 해당하는 구간이 없습니다.")

                if ext_errors:
                    for ext_error in ext_errors:
                        st.error(ext_error)
                else:
                    evidence_path = str(external_row.get("외부증빙파일", "") or "")
                    if external_evidence is not None:
                        evidence_name = (
                            f"external_{st.session_state.current_candidate_id or 'temp'}_"
                            f"{safe_filename(external_company)}_{datetime.now():%Y%m%d_%H%M%S}_"
                            f"{safe_filename(external_evidence.name)}"
                        )
                        evidence_target = UPLOAD_DIR / evidence_name
                        evidence_target.write_bytes(external_evidence.getvalue())
                        evidence_path = str(evidence_target)

                    ext_rate = float(ext_rule["recognition_rate"])
                    career_for_external.at[selected_external_idx, "표준회사명"] = str(external_row.get("표준회사명", "") or external_company)
                    career_for_external.at[selected_external_idx, "매출연도"] = int(external_year)
                    career_for_external.at[selected_external_idx, "연결매출(억원)"] = round(float(external_revenue), 2)
                    career_for_external.at[selected_external_idx, "매출조회상태"] = "외부자료확정"
                    career_for_external.at[selected_external_idx, "매출기준인정률(%)"] = ext_rate
                    career_for_external.at[selected_external_idx, "공시접수번호"] = ""
                    career_for_external.at[selected_external_idx, "외부자료출처"] = external_source_type
                    career_for_external.at[selected_external_idx, "외부자료URL"] = external_url.strip()
                    career_for_external.at[selected_external_idx, "재무기준"] = external_statement_type
                    career_for_external.at[selected_external_idx, "외부증빙파일"] = evidence_path
                    career_for_external.at[selected_external_idx, "수동확정"] = True
                    career_for_external.at[selected_external_idx, "수동조정사유"] = external_reason.strip() or f"{external_year}년 외부자료 연결매출 확인"
                    career_for_external.at[selected_external_idx, "인정구분"] = "수동확정"
                    career_for_external.at[selected_external_idx, "인정률(%)"] = ext_rate
                    career_for_external.at[selected_external_idx, "출처"] = "외부자료 수동확정"

                    st.session_state.career_df = ensure_career_columns(career_for_external)
                    st.session_state.career_editor_version += 1
                    if st.session_state.current_candidate_id:
                        replace_career_records(
                            st.session_state.current_candidate_id,
                            career_records_for_db(st.session_state.career_df),
                        )
                        add_audit_log(
                            st.session_state.current_candidate_id,
                            "외부자료 연결매출 수동확정",
                            {
                                "회사명": external_company,
                                "사업연도": int(external_year),
                                "연결매출_억원": float(external_revenue),
                                "출처": external_source_type,
                                "URL": external_url.strip(),
                                "인정률": ext_rate,
                            },
                        )
                    st.success(f"{external_company}: {external_revenue:,.1f}억원 → 인정률 {ext_rate:.0f}%로 확정했습니다.")
                    st.rerun()
        else:
            st.success("현재 외부자료 확인이 필요한 경력회사가 없습니다.")

        career_result = calculate_career(st.session_state.career_df, calculation_date)
        for error in career_result["errors"]:
            st.error(error)

        metric1, metric2, metric3, metric4 = st.columns(4)
        metric1.metric("총 실경력", period_text(career_result["raw_days"]))
        metric2.metric("확정 인정경력", period_text(career_result["recognized_days"]))
        metric3.metric("확인필요 경력", period_text(career_result["pending_days"]))
        ratio = (
            career_result["recognized_days"] / career_result["raw_days"] * 100
            if career_result["raw_days"] else 0
        )
        metric4.metric("확정 경력 인정비율", f"{ratio:,.1f}%")

        if not career_result["detail"].empty:
            st.dataframe(career_result["detail"], hide_index=True, use_container_width=True)

    with tabs[3]:
        st.subheader("현재 보상과 당사 제안 비교")

        current_col, band_col = st.columns([1, 1.4])
        with current_col:
            st.markdown("#### 현재 보상")
            money_input("계약연봉(원)", "c_current_contract_salary", step=100000)
            money_input("총연봉(원) (원천징수 고려)", "c_current_total_salary", step=100000)

        with band_col:
            st.markdown("#### Pay Band")
            grade_col, role_col, band_select_col = st.columns(3)
            with grade_col:
                st.selectbox("직급", GRADE_OPTIONS, key="c_grade")
            with role_col:
                role_options = ["일반", "팀장", "실장"] if st.session_state.c_grade in {"G4", "R4"} else ["일반"]
                if st.session_state.c_grade_role not in role_options:
                    st.session_state.c_grade_role = "일반"
                st.selectbox("직책구분", role_options, key="c_grade_role")

            grade_group = pay_grade_group(st.session_state.c_grade, st.session_state.c_grade_role)
            band_reference = pd.DataFrame(get_pay_band_reference())
            available_bands = []
            if not band_reference.empty:
                available_bands = band_reference.loc[band_reference["직급"] == grade_group, "BAND"].astype(str).tolist()
            available_bands = list(dict.fromkeys([b for b in available_bands if b and b != "nan"]))
            with band_select_col:
                if available_bands:
                    if st.session_state.c_band not in available_bands:
                        st.session_state.c_band = available_bands[0]
                    st.selectbox("BAND", available_bands, key="c_band")
                else:
                    st.selectbox("BAND", ["관리자 기준 없음"], key="_band_no_reference", disabled=True)
                    st.session_state.c_band = ""

            selected_key = f"{grade_group}|{st.session_state.c_band}" if st.session_state.c_band else ""
            selected_row = get_pay_band_row(grade_group, st.session_state.c_band) if st.session_state.c_band else None
            if selected_row and selected_key != st.session_state.c_last_band_key:
                st.session_state.c_offer_base_salary = int(round(float(selected_row.get("기본연봉", 0) or 0)))
                st.session_state.c_offer_performance_salary = int(round(float(selected_row.get("업적연봉", 0) or 0)))
                st.session_state.c_offer_fixed_overtime = int(round(float(selected_row.get("고정연장", 0) or 0)))
                st.session_state.c_offer_incentive = int(round(float(selected_row.get("경영성과급", 0) or 0)))
                st.session_state.c_cash_job_allowance = int(round(float(selected_row.get("직책수당", 0) or 0)))
                st.session_state.c_cash_family_allowance = int(round(float(selected_row.get("가족수당(본인)", 0) or 0)))
                st.session_state.c_cash_vehicle_subsidy = int(round(float(selected_row.get("차량보조금", 0) or 0)))
                st.session_state.c_cash_self_development = int(round(float(selected_row.get("자기계발지원금", 0) or 0)))
                st.session_state.c_cash_homecoming_travel = int(round(float(selected_row.get("귀향여비", 0) or 0)))
                st.session_state.c_cash_welfare_points = int(round(float(selected_row.get("복지포인트", 0) or 0)))
                st.session_state.c_cash_flexible_welfare_points = int(round(float(selected_row.get("선택형 복지포인트", 0) or 0)))
                st.session_state.c_cash_fitness = int(round(float(selected_row.get("체력단련비", 0) or 0)))
                st.session_state.c_cash_pension = int(round(float(selected_row.get("개인연금", 0) or 0)))
                st.session_state.c_last_band_key = selected_key

            st.text_input("승급기준일", key="c_promotion_base_date", placeholder="예: 26.01")

            if selected_row:
                st.caption(f"관리자 기준: {grade_group} / {st.session_state.c_band} · 기준 총계 {format_won(selected_row.get('총계', 0))}")
                if st.button("현재 Pay Band 기준 다시 불러오기", key="reload_selected_pay_band"):
                    st.session_state.c_offer_base_salary = int(round(float(selected_row.get("기본연봉", 0) or 0)))
                    st.session_state.c_offer_performance_salary = int(round(float(selected_row.get("업적연봉", 0) or 0)))
                    st.session_state.c_offer_fixed_overtime = int(round(float(selected_row.get("고정연장", 0) or 0)))
                    st.session_state.c_offer_incentive = int(round(float(selected_row.get("경영성과급", 0) or 0)))
                    st.session_state.c_cash_job_allowance = int(round(float(selected_row.get("직책수당", 0) or 0)))
                    st.session_state.c_cash_family_allowance = int(round(float(selected_row.get("가족수당(본인)", 0) or 0)))
                    st.session_state.c_cash_vehicle_subsidy = int(round(float(selected_row.get("차량보조금", 0) or 0)))
                    st.session_state.c_cash_self_development = int(round(float(selected_row.get("자기계발지원금", 0) or 0)))
                    st.session_state.c_cash_homecoming_travel = int(round(float(selected_row.get("귀향여비", 0) or 0)))
                    st.session_state.c_cash_welfare_points = int(round(float(selected_row.get("복지포인트", 0) or 0)))
                    st.session_state.c_cash_flexible_welfare_points = int(round(float(selected_row.get("선택형 복지포인트", 0) or 0)))
                    st.session_state.c_cash_fitness = int(round(float(selected_row.get("체력단련비", 0) or 0)))
                    st.session_state.c_cash_pension = int(round(float(selected_row.get("개인연금", 0) or 0)))
                    st.session_state.c_last_band_key = selected_key
                    st.rerun()
            else:
                st.warning("선택한 직급의 Pay Band 기준이 없습니다. 관리자 설정에서 CSV 업로드 또는 표 입력 후 저장하세요.")

        st.markdown("#### 당사 제안")
        st.markdown("##### 계약연봉")
        st.caption("계약연봉 소계 = 기본급 + 업적급 + 고정연장수당. 성과급은 계약연봉에는 포함하지 않고 총연봉 비교에 별도 포함합니다.")
        p1, p2, p3, p4 = st.columns(4)
        with p1:
            money_input("제안 기본급(원)", "c_offer_base_salary", step=100000)
        with p2:
            money_input("제안 업적급(원)", "c_offer_performance_salary", step=100000)
        with p3:
            money_input("제안 고정연장수당(원)", "c_offer_fixed_overtime", step=100000)
        with p4:
            money_input("성과급(원)", "c_offer_incentive", step=100000, help="경영성과급 등 계약연봉 외 성과급")

        st.markdown("##### 현금성지급 복리후생")
        cash_cols = st.columns(3)
        for idx, (label, key) in enumerate(CASH_BENEFIT_FIELDS):
            with cash_cols[idx % 3]:
                money_input(f"{label}(원)", key, step=10000)

        comp = compensation_result()
        summary1, summary2, summary3 = st.columns(3)
        summary1.metric("당사 제안 계약연봉", format_won(comp["offered_contract"]))
        summary2.metric("현금성지급 복리후생 소계", format_won(comp["cash_benefits"]))
        summary3.metric("당사 제안 총연봉", format_won(comp["offered_total"]))

        st.markdown("#### 처우 비교")
        left, right = st.columns(2)
        with left:
            st.markdown("##### 계약연봉 기준")
            c1, c2, c3 = st.columns(3)
            c1.metric("현재 계약연봉", format_won(comp["current_contract_salary"]))
            c2.metric("당사 계약연봉", format_won(comp["offered_contract"]))
            c3.metric(
                f"계약연봉 {comp['contract_result']}",
                format_won(abs(comp["contract_difference"])),
                delta=format_percent(comp["contract_difference_rate"]),
            )
        with right:
            st.markdown("##### 총연봉 기준")
            t1, t2, t3 = st.columns(3)
            t1.metric("현재 총연봉", format_won(comp["current_total_salary"]))
            t2.metric("당사 총연봉", format_won(comp["offered_total"]))
            t3.metric(
                f"총연봉 {comp['result']}",
                format_won(abs(comp["difference"])),
                delta=format_percent(comp["difference_rate"]),
            )

        st.caption("총연봉 비교에는 계약연봉 + 성과급 + 현금성지급 복리후생을 포함합니다. 현재 총연봉은 원천징수 자료 등을 고려해 입력하는 값입니다.")

    with tabs[4]:
        st.subheader("입사신분별 복리후생")
        profile_label = (
            f"{st.session_state.f_job_group} / {st.session_state.f_employment_type} / "
            f"{st.session_state.f_work_location} / {st.session_state.f_entry_type} / {st.session_state.c_grade}"
        )
        profile_key = benefit_profile_key(
            st.session_state.f_job_group, st.session_state.f_employment_type,
            st.session_state.f_work_location, st.session_state.f_entry_type, st.session_state.c_grade,
        )
        st.caption(f"현재 적용 기준: {profile_label}")

        col_load, col_info = st.columns([1, 3])
        with col_load:
            if st.button("관리자 복리후생 기준 불러오기", use_container_width=True):
                loaded_benefits = benefit_profile_df(
                    st.session_state.f_job_group, st.session_state.f_employment_type,
                    st.session_state.f_work_location, st.session_state.f_entry_type, st.session_state.c_grade,
                )
                st.session_state.benefit_df = loaded_benefits
                st.session_state.benefit_editor_version += 1
                st.rerun()
        with col_info:
            if profile_key not in benefit_profiles():
                st.warning("현재 조건의 관리자 복리후생 기준이 아직 저장되지 않았습니다. 관리자 설정에서 먼저 체크하여 저장하거나 이 화면에서 직접 항목을 추가하세요.")
            else:
                st.info("관리자 기준을 불러온 뒤에도 아래 표에서 후보자별로 복리후생을 추가·삭제하거나 설명을 수정할 수 있습니다.")

        benefit_source = st.session_state.benefit_df.copy()
        for col in BENEFIT_COLUMNS:
            if col not in benefit_source.columns:
                benefit_source[col] = False if col == "오퍼레터 포함" else ""
        benefit_source = benefit_source[BENEFIT_COLUMNS]
        benefit_editor = st.data_editor(
            benefit_source,
            hide_index=True,
            use_container_width=True,
            num_rows="dynamic",
            key=f"benefit_editor_{st.session_state.benefit_editor_version}",
            column_config={
                "오퍼레터 포함": st.column_config.CheckboxColumn("오퍼레터 포함", default=True),
                "구분": st.column_config.TextColumn("구분", width="small"),
                "복리후생": st.column_config.TextColumn("복리후생", width="medium"),
                "적용여부": st.column_config.SelectboxColumn(
                    "적용여부", options=["적용", "조건부", "미적용"], default="적용",
                ),
                "설명": st.column_config.TextColumn("설명", width="large"),
            },
        )
        benefit_editor = benefit_editor[
            benefit_editor["복리후생"].fillna("").astype(str).str.strip() != ""
        ].copy()
        st.session_state.benefit_df = benefit_editor

        col1, col2, col3 = st.columns(3)
        col1.metric("적용", f"{(benefit_editor['적용여부'] == '적용').sum()}개")
        col2.metric("조건부", f"{(benefit_editor['적용여부'] == '조건부').sum()}개")
        col3.metric("미적용", f"{(benefit_editor['적용여부'] == '미적용').sum()}개")

    with tabs[5]:
        st.subheader("오퍼레터 PPT 생성 및 버전관리")
        col1, col2 = st.columns(2)
        with col1:
            st.date_input("오퍼 발행일", key="offer_date")
        with col2:
            st.date_input("수락 회신기한", key="acceptance_deadline")
        st.text_area(
            "추가 조건 또는 안내문구",
            key="special_terms",
            height=120,
            placeholder="예: 사이닝 보너스 반환조건, 수습기간, 별도 승인조건 등",
        )

        career_result = calculate_career(st.session_state.career_df, date.today())
        comp = compensation_result()
        selected_benefits = st.session_state.benefit_df[
            st.session_state.benefit_df["오퍼레터 포함"] == True  # noqa: E712
        ].to_dict("records")

        summary_df = pd.DataFrame([
            {"항목": "후보자", "내용": st.session_state.f_name or "미입력"},
            {"항목": "입사신분", "내용": st.session_state.f_employment_type},
            {"항목": "지원직무", "내용": st.session_state.f_target_job or "미입력"},
            {"항목": "인정경력", "내용": period_text(career_result["recognized_days"])},
            {"항목": "제안 계약연봉", "내용": format_won(comp["offered_contract"])},
            {"항목": "제안 총연봉", "내용": format_won(comp["offered_total"])},
            {"항목": "계약연봉 비교", "내용": f"{comp['contract_result']} ({format_percent(comp['contract_difference_rate'])})"},
            {"항목": "총연봉 비교", "내용": f"{comp['result']} ({format_percent(comp['difference_rate'])})"},
            {"항목": "포함 복리후생", "내용": f"{len(selected_benefits)}개"},
        ])
        st.dataframe(summary_df, hide_index=True, use_container_width=True)

        if st.button("오퍼레터 생성·저장", type="primary"):
            candidate_id = save_current_candidate(
                reason="오퍼레터 생성 전 저장",
                show_message=False,
            )
            if candidate_id:
                try:
                    document_bytes = build_offer_ppt(
                        company_name=COMPANY_NAME,
                        hr_department=HR_DEPARTMENT,
                        candidate_name=st.session_state.f_name,
                        job_title=st.session_state.f_target_job,
                        department=st.session_state.f_department,
                        work_location=st.session_state.f_work_location,
                        employment_type=st.session_state.f_employment_type,
                        expected_join_date=st.session_state.f_expected_join_date,
                        offer_date=st.session_state.offer_date,
                        acceptance_deadline=st.session_state.acceptance_deadline,
                        recognized_career_text=period_text(career_result["recognized_days"]),
                        base_salary=comp["offer_base_salary"],
                        performance_salary=comp["offer_performance_salary"],
                        fixed_overtime=comp["offer_fixed_overtime"],
                        incentive=comp["offer_incentive"],
                        cash_benefit_values={
                            "직책수당": comp["cash_job_allowance"],
                            "가족수당(본인)": comp["cash_family_allowance"],
                            "차량보조금": comp["cash_vehicle_subsidy"],
                            "자기계발지원금": comp["cash_self_development"],
                            "귀향여비": comp["cash_homecoming_travel"],
                            "복지포인트": comp["cash_welfare_points"],
                            "선택형 복지포인트": comp["cash_flexible_welfare_points"],
                            "체력단련비": comp["cash_fitness"],
                            "개인연금": comp["cash_pension"],
                        },
                        sign_on_bonus=comp["sign_on_bonus"],
                        benefits=selected_benefits,
                        special_terms=st.session_state.special_terms,
                    )

                    version_no = next_offer_version(candidate_id)
                    offer_dir = OFFER_DIR / safe_filename(st.session_state.candidate_code)
                    offer_dir.mkdir(parents=True, exist_ok=True)
                    file_name = (
                        f"Offer_{safe_filename(st.session_state.f_name)}_"
                        f"v{version_no}_{date.today():%Y%m%d}.pptx"
                    )
                    file_path = offer_dir / file_name
                    file_path.write_bytes(document_bytes)

                    snapshot = build_snapshot()
                    add_offer_version(
                        candidate_id,
                        version_no,
                        str(file_path),
                        snapshot,
                    )
                    add_audit_log(
                        candidate_id,
                        "오퍼레터 생성",
                        {"version": version_no, "file_name": file_name},
                    )

                    st.success(f"오퍼레터 v{version_no}을 저장했습니다.")
                    st.download_button(
                        "방금 생성한 오퍼레터 PPT 다운로드",
                        data=document_bytes,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    )
                except Exception as exc:
                    st.error(f"오퍼레터 생성 중 오류: {exc}")

        if st.session_state.current_candidate_id:
            versions = get_offer_versions(st.session_state.current_candidate_id)
            if versions:
                version_df = pd.DataFrame(versions)
                st.markdown("#### 기존 오퍼레터 버전")
                st.dataframe(
                    version_df[["version_no", "created_at", "file_path"]].rename(columns={
                        "version_no": "버전",
                        "created_at": "생성일시",
                        "file_path": "저장경로",
                    }),
                    hide_index=True,
                    use_container_width=True,
                )
                version_map = {
                    f"v{item['version_no']} | {item['created_at']}": item
                    for item in versions
                }
                selected_version_label = st.selectbox(
                    "다운로드할 버전",
                    list(version_map.keys()),
                )
                selected_version = version_map[selected_version_label]
                version_path = Path(selected_version["file_path"])
                if version_path.exists():
                    st.download_button(
                        "선택 버전 다운로드",
                        data=version_path.read_bytes(),
                        file_name=version_path.name,
                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    )

    with tabs[6]:
        st.subheader("저장·변경 이력")
        if not st.session_state.current_candidate_id:
            st.info("후보자를 저장하면 변경이력이 표시됩니다.")
        else:
            logs = get_audit_logs(st.session_state.current_candidate_id)
            if logs:
                log_df = pd.DataFrame(logs)
                st.dataframe(
                    log_df[["created_at", "action", "detail_json"]].rename(columns={
                        "created_at": "일시",
                        "action": "작업",
                        "detail_json": "세부내용",
                    }),
                    hide_index=True,
                    use_container_width=True,
                )

            snapshots = get_snapshots(st.session_state.current_candidate_id)
            if snapshots:
                snapshot_map = {
                    f"{item['created_at']} | {item.get('reason', '')}": item
                    for item in snapshots
                }
                selected_snapshot_label = st.selectbox(
                    "과거 저장 스냅샷 확인",
                    list(snapshot_map.keys()),
                )
                selected_snapshot = snapshot_map[selected_snapshot_label]
                with st.expander("선택 스냅샷 상세", expanded=False):
                    parsed = json.loads(selected_snapshot["snapshot_json"])
                    st.json(parsed)


elif nav == "관리자 설정":
    st.title("관리자 설정")
    st.warning(
        "현재 버전에는 로그인·권한통제가 없습니다. 실제 운영 시 관리자 권한을 별도로 적용하세요."
    )
    admin_tabs = st.tabs(["DART 연동", "매출 인정기준", "Pay Band 기준", "복리후생 기준", "회사명 매칭사전"])

    with admin_tabs[0]:
        st.subheader("OpenDART 연동")
        saved_key = get_setting("DART_API_KEY", "")
        dart_key = st.text_input(
            "DART API 인증키",
            value=saved_key,
            type="password",
            max_chars=40,
            help="로컬 프로토타입에서는 data/offer_system.db에 저장됩니다.",
        )
        col1, col2 = st.columns(2)
        with col1:
            if st.button("API 키 저장", type="primary", use_container_width=True):
                if len(dart_key.strip()) != 40:
                    st.error("DART API 인증키는 40자리인지 확인하세요.")
                else:
                    set_setting("DART_API_KEY", dart_key.strip())
                    st.success("DART API 키를 저장했습니다.")
        with col2:
            if st.button("DART 기업목록 새로고침", use_container_width=True):
                key_to_use = dart_key.strip() or saved_key.strip()
                try:
                    with st.spinner("DART 전체 기업목록을 내려받고 있습니다."):
                        count = download_corp_codes(key_to_use, DART_CORP_CACHE)
                    st.success(f"DART 기업목록 {count:,}개를 저장했습니다.")
                except Exception as exc:
                    st.error(f"기업목록 갱신 실패: {exc}")
        if DART_CORP_CACHE.exists():
            st.caption(
                f"기업목록 파일: {DART_CORP_CACHE.name} · "
                f"최종수정 {datetime.fromtimestamp(DART_CORP_CACHE.stat().st_mtime):%Y-%m-%d %H:%M:%S}"
            )
        else:
            st.info("DART 기업목록을 한 번 새로고침하세요.")

    with admin_tabs[1]:
        st.subheader("연결매출 구간별 경력 인정률")
        st.info(
            "단위는 억원입니다. 하한 이상·상한 미만으로 적용하며, 최상위 구간은 상한을 비워둡니다. "
            "어떤 구간에도 포함되지 않으면 확인필요로 처리합니다. 최초 표시 기준은 예시이므로 실제 회사 기준으로 수정하세요."
        )
        rules = get_revenue_rules()
        rules_df = pd.DataFrame([
            {
                "구간명": item.get("label", ""),
                "매출하한(억원)": item.get("min_revenue_eok", 0),
                "매출상한(억원)": item.get("max_revenue_eok"),
                "인정률(%)": item.get("recognition_rate", 0),
            }
            for item in rules
        ])
        edited_rules = st.data_editor(
            rules_df,
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "구간명": st.column_config.TextColumn("구간명"),
                "매출하한(억원)": st.column_config.NumberColumn(
                    "매출하한(억원)", min_value=0.0, step=100.0, format="%.0f",
                ),
                "매출상한(억원)": st.column_config.NumberColumn(
                    "매출상한(억원)", min_value=0.0, step=100.0, format="%.0f",
                ),
                "인정률(%)": st.column_config.NumberColumn(
                    "인정률(%)", min_value=0.0, max_value=100.0, step=5.0, format="%.0f%%",
                ),
            },
        )
        if st.button("매출 인정기준 저장", type="primary"):
            errors = []
            clean_records = []
            for row_no, row in enumerate(edited_rules.to_dict("records"), start=1):
                minimum = float(row.get("매출하한(억원)", 0) or 0)
                maximum_raw = row.get("매출상한(억원)")
                maximum = None if pd.isna(maximum_raw) or maximum_raw == "" else float(maximum_raw)
                rate = float(row.get("인정률(%)", 0) or 0)
                if maximum is not None and maximum <= minimum:
                    errors.append(f"{row_no}행: 상한은 하한보다 커야 합니다.")
                clean_records.append({**row, "매출하한(억원)": minimum, "매출상한(억원)": maximum, "인정률(%)": rate})

            sorted_rows = sorted(clean_records, key=lambda item: item["매출하한(억원)"])
            for left, right in zip(sorted_rows, sorted_rows[1:]):
                left_max = left["매출상한(억원)"]
                if left_max is None or left_max > right["매출하한(억원)"]:
                    errors.append("매출 구간이 서로 겹치거나 중간 상한이 비어 있습니다.")
                    break
            if not clean_records:
                errors.append("최소 1개 매출 구간이 필요합니다.")

            if errors:
                for error in errors:
                    st.error(error)
            else:
                replace_revenue_rules(clean_records)
                st.success("매출 인정기준을 저장했습니다. 이후 일괄조회부터 적용됩니다.")

    with admin_tabs[2]:
        st.subheader("Pay Band 기준")
        st.info(
            "단위는 원입니다. CSV 업로드 또는 아래 표에 Excel 범위를 그대로 붙여넣어 수정할 수 있습니다. "
            "저장한 값은 후보자 보상비교의 직급/BAND 선택과 동시에 당사 제안 금액으로 불러옵니다."
        )
        uploaded_pay_csv = st.file_uploader("Pay Band CSV 업로드", type=["csv"], key="pay_band_csv_upload")
        saved_pay_df = normalize_pay_band_dataframe(pd.DataFrame(get_pay_band_reference()))
        if uploaded_pay_csv is not None:
            try:
                pay_edit_source = normalize_pay_band_dataframe(parse_pay_band_csv(uploaded_pay_csv))
                st.success(f"CSV {len(pay_edit_source):,}행을 불러왔습니다. 아래 표 확인 후 저장하세요.")
            except Exception as exc:
                st.error(str(exc))
                pay_edit_source = saved_pay_df
        else:
            pay_edit_source = saved_pay_df

        pay_editor = st.data_editor(
            pay_edit_source,
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            key="pay_band_reference_editor",
            column_config={
                "직급": st.column_config.TextColumn("직급", width="medium"),
                "BAND": st.column_config.TextColumn("BAND", width="medium"),
                **{col: st.column_config.NumberColumn(col, min_value=0, step=10000, format="%,d") for col in PAY_BAND_COLUMNS[2:]},
            },
        )
        st.caption("권장 직급값: G4/R4(실장), G4/R4(팀장), G4/R4, G3/R3, G2/R2, G1/R1")
        if st.button("Pay Band 기준 저장", type="primary"):
            cleaned = normalize_pay_band_dataframe(pay_editor)
            cleaned = cleaned[(cleaned["직급"] != "") & (cleaned["BAND"] != "")].copy()
            duplicate = cleaned.duplicated(subset=["직급", "BAND"], keep=False)
            if cleaned.empty:
                st.error("저장할 Pay Band 행이 없습니다.")
            elif duplicate.any():
                st.error("직급+BAND 조합이 중복되어 있습니다. 중복 행을 정리하세요.")
            else:
                # 소계/총계가 비어 있거나 0이어도 입력 구성요소 기준으로 자동 보정합니다.
                cleaned["계약연봉소계"] = cleaned["기본연봉"] + cleaned["업적연봉"] + cleaned["고정연장"]
                cash_cols = ["직책수당", "가족수당(본인)", "차량보조금", "자기계발지원금", "귀향여비", "복지포인트", "선택형 복지포인트", "체력단련비", "개인연금"]
                cleaned["현금성지급소계"] = cleaned[cash_cols].sum(axis=1)
                cleaned["계약연봉+현금성지급"] = cleaned["계약연봉소계"] + cleaned["현금성지급소계"]
                cleaned["총계"] = cleaned["계약연봉+현금성지급"] + cleaned["경영성과급"]
                replace_pay_band_reference(cleaned.to_dict("records"))
                st.success(f"Pay Band 기준 {len(cleaned):,}행을 저장했습니다.")
                st.rerun()

        st.download_button(
            "현재 Pay Band 기준 CSV 다운로드",
            data=normalize_pay_band_dataframe(pd.DataFrame(get_pay_band_reference())).to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig"),
            file_name="pay_band_reference.csv",
            mime="text/csv",
        )

    with admin_tabs[3]:
        st.subheader("입사신분별 복리후생 기준")
        st.caption(
            "직군·계약구분·근무지·입사구분·직급 조합별로 적용할 복리후생을 체크하여 저장합니다. "
            "후보자 화면에서는 이 기준을 불러온 뒤 개별 추가·삭제가 가능합니다."
        )

        with st.expander("복리후생 항목 마스터 및 설명 편집", expanded=False):
            st.info(
                "기본 항목은 현대제철 채용 Benefits 및 공개적으로 확인 가능한 복리후생 범주를 기준으로 구성했습니다. "
                "귀향여비·선택형 복지포인트·자기계발지원금·체력단련비·개인연금은 현금성 보상에서 관리하므로 제외했습니다."
            )
            catalog_editor = st.data_editor(
                benefit_catalog_df(),
                num_rows="dynamic",
                hide_index=True,
                use_container_width=True,
                key="benefit_catalog_editor",
                column_config={
                    "사용": st.column_config.CheckboxColumn("사용", default=True),
                    "구분": st.column_config.TextColumn("구분", width="small"),
                    "복리후생": st.column_config.TextColumn("복리후생", width="medium"),
                    "설명": st.column_config.TextColumn("설명", width="large"),
                },
            )
            if st.button("복리후생 항목 마스터 저장", type="primary"):
                save_benefit_catalog(catalog_editor)
                st.success("복리후생 항목과 설명을 저장했습니다.")
                st.rerun()

        b1, b2, b3, b4, b5 = st.columns(5)
        with b1:
            admin_job_group = st.selectbox("직군", JOB_GROUP_OPTIONS, key="admin_benefit_job_group")
        with b2:
            admin_contract_type = st.selectbox("계약구분", EMPLOYMENT_TYPES, key="admin_benefit_contract_type")
        with b3:
            admin_location = st.selectbox("근무지", WORK_LOCATION_OPTIONS, key="admin_benefit_location")
        with b4:
            admin_entry_type = st.selectbox("입사구분", ENTRY_TYPE_OPTIONS, key="admin_benefit_entry_type")
        with b5:
            admin_grade = st.selectbox("직급", GRADE_OPTIONS, key="admin_benefit_grade")

        admin_profile_key = benefit_profile_key(
            admin_job_group, admin_contract_type, admin_location, admin_entry_type, admin_grade,
        )
        all_profiles = benefit_profiles()
        selected_names = set(all_profiles.get(admin_profile_key, []))
        active_catalog = benefit_catalog_df()
        active_catalog = active_catalog[active_catalog["사용"] == True].copy()  # noqa: E712
        active_catalog.insert(0, "적용", active_catalog["복리후생"].isin(selected_names))
        st.markdown(f"##### 적용항목 체크 · {admin_job_group} / {admin_contract_type} / {admin_location} / {admin_entry_type} / {admin_grade}")
        profile_editor = st.data_editor(
            active_catalog[["적용", "구분", "복리후생", "설명"]],
            hide_index=True,
            use_container_width=True,
            key=f"benefit_profile_editor_{admin_profile_key}",
            column_config={
                "적용": st.column_config.CheckboxColumn("적용"),
                "구분": st.column_config.TextColumn("구분", width="small"),
                "복리후생": st.column_config.TextColumn("복리후생", width="medium"),
                "설명": st.column_config.TextColumn("설명", width="large"),
            },
            disabled=["구분", "복리후생", "설명"],
        )
        save_col, delete_col, stat_col = st.columns([1, 1, 2])
        with save_col:
            if st.button("현재 조건 복리후생 저장", type="primary", use_container_width=True):
                all_profiles[admin_profile_key] = profile_editor.loc[
                    profile_editor["적용"] == True, "복리후생"  # noqa: E712
                ].astype(str).tolist()
                save_benefit_profiles(all_profiles)
                st.success(f"복리후생 {len(all_profiles[admin_profile_key])}개를 저장했습니다.")
                st.rerun()
        with delete_col:
            if st.button("현재 조건 기준 삭제", use_container_width=True):
                all_profiles.pop(admin_profile_key, None)
                save_benefit_profiles(all_profiles)
                st.success("현재 조건의 복리후생 기준을 삭제했습니다.")
                st.rerun()
        with stat_col:
            st.metric("저장된 입사조건 프로필", f"{len(all_profiles):,}개")

    with admin_tabs[4]:
        st.subheader("건강보험 회사명 ↔ DART 회사 매칭사전")
        st.caption(
            "자동매칭이 불확실한 회사는 여기에서 DART 회사를 검색해 직접 매칭하거나, "
            "경력표에 DART 고유번호를 입력한 뒤 다시 조회할 수 있습니다."
        )

        if DART_CORP_CACHE.exists():
            corp_df = load_corp_codes(DART_CORP_CACHE)
            search_corp = st.text_input("DART 기업명 검색", placeholder="회사명 일부 입력")
            if search_corp.strip():
                normalized_query = normalize_company_name(search_corp)
                search_result = corp_df[
                    corp_df["normalized_name"].str.contains(normalized_query, na=False, regex=False)
                    | corp_df["corp_name"].str.contains(search_corp.strip(), na=False, regex=False)
                ].head(50)
                st.dataframe(
                    search_result[["corp_code", "corp_name", "stock_code"]].rename(columns={
                        "corp_code": "DART고유번호",
                        "corp_name": "DART회사명",
                        "stock_code": "종목코드",
                    }),
                    hide_index=True,
                    use_container_width=True,
                )
        else:
            st.info("DART 연동 탭에서 기업목록을 먼저 새로고침하세요.")

        st.markdown("#### 수동 매칭 등록")
        map_col1, map_col2, map_col3 = st.columns(3)
        with map_col1:
            mapping_alias = st.text_input("건강보험상 회사명")
        with map_col2:
            mapping_code = st.text_input("DART 고유번호", max_chars=8)
            normalized_mapping_code = normalize_corp_code(mapping_code)
        with map_col3:
            mapping_name = st.text_input("DART 공식 회사명")
        if st.button("수동 매칭 저장"):
            if not mapping_alias.strip() or len(mapping_code.strip()) != 8 or not mapping_name.strip():
                st.error("회사명, 8자리 DART 고유번호, DART 공식 회사명을 모두 입력하세요.")
            else:
                upsert_company_mapping(
                    mapping_alias.strip(),
                    normalize_company_name(mapping_alias),
                    normalized_mapping_code,
                    mapping_name.strip(),
                )
                st.success("회사명 매칭을 저장했습니다.")
                st.rerun()

        st.markdown("#### 저장된 매칭")
        mappings = list_company_mappings()
        if mappings:
            mapping_df = pd.DataFrame(mappings).rename(columns={
                "alias_name": "건강보험상 회사명",
                "corp_code": "DART고유번호",
                "corp_name": "DART회사명",
                "updated_at": "최종수정",
            })
            st.dataframe(mapping_df, hide_index=True, use_container_width=True)
            delete_alias = st.selectbox(
                "삭제할 매칭",
                [item["alias_name"] for item in mappings],
            )
            if st.button("선택 매칭 삭제"):
                delete_company_mapping(delete_alias)
                st.success("매칭을 삭제했습니다.")
                st.rerun()
        else:
            st.info("저장된 회사명 매칭이 없습니다.")


else:
    st.title("시스템 안내")
    st.markdown(
        """
        ### 현재 제공 기능

        - 후보자 여러 명 등록 및 검색
        - 후보자별 기본정보·상태 저장
        - 건강보험 자격득실확인서 PDF·이미지 업로드
        - 문자형 PDF 직접 추출 및 스캔본 OCR
        - OCR 경력 후보를 인정경력 표로 자동 추가
        - 중복 경력 제거 및 인정률 반영
        - 현재 보상과 당사 제안 보상 비교
        - 입사신분별 복리후생 검토
        - 오퍼레터 PPT 생성 및 버전관리
        - 저장 스냅샷과 작업이력 조회

        ### 데이터 저장 위치

        프로그램 폴더 아래 `data` 폴더에 저장됩니다.

        - DB: `data/offer_system.db`
        - 업로드 문서: `data/uploads`
        - 오퍼레터: `data/offers`

        ### 운영 전 필수 보완

        이 버전은 로컬 프로토타입입니다. 실제 지원자 개인정보와 연봉자료를 처리하려면
        사내 승인 서버, 로그인, 권한관리, 암호화, 접근로그, 백업, 보존기간과 파기정책을
        별도로 적용해야 합니다.

        건강보험 자격득실확인서만으로 실제 직무·직급·고용형태·휴직기간을 확정할 수 없습니다.
        OCR 결과와 인정률은 경력증명서, 이력서 및 채용부서 확인을 거쳐 담당자가 확정해야 합니다.
        """
    )

    st.markdown("### OCR 점검")
    st.json(tesseract_status())
